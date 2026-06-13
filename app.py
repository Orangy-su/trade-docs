"""
app_v3.py — 外贸单据系统 v3
- 主数据手册固化（上传一次长期使用）
- 网页直接选客户/卖方，填Invoice No.，定义货柜分组
- 新客户网页录入，自动写入主数据手册
- ⑤出货批次记录由网页表单驱动，无需手动维护
"""
import os, sys, json, uuid, threading, traceback, io
from flask import Flask, request, jsonify, send_file, render_template_string
from datetime import datetime

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
ENGINE_DIR = os.path.join(BASE_DIR, 'engine')
sys.path.insert(0, BASE_DIR)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
app.secret_key = os.environ.get('SECRET_KEY', 'qf-trade-v3-2026')

# 云端部署用 /tmp（可写），本地用项目目录
IS_CLOUD = (os.environ.get('RAILWAY_ENVIRONMENT') is not None or
            os.environ.get('RENDER') is not None)
STORE_BASE = '/tmp' if IS_CLOUD else BASE_DIR

UPLOAD_DIR  = os.path.join(STORE_BASE, 'uploads');    os.makedirs(UPLOAD_DIR, exist_ok=True)
OUTPUT_DIR  = os.path.join(STORE_BASE, 'output');     os.makedirs(OUTPUT_DIR, exist_ok=True)
MASTER_DIR  = os.path.join(STORE_BASE, 'master_store'); os.makedirs(MASTER_DIR, exist_ok=True)
MASTER_META = os.path.join(MASTER_DIR, 'meta.json')

tasks = {}

def load_master_meta():
    if os.path.exists(MASTER_META):
        with open(MASTER_META) as f: return json.load(f)
    return None

def save_master_meta(meta):
    with open(MASTER_META, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def get_master_path():
    meta = load_master_meta()
    if meta and os.path.exists(meta.get('path','')):
        return meta['path']
    return None


HTML = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>外贸单据系统 · 擎烽</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;700&family=JetBrains+Mono:wght@400;700&display=swap" rel="stylesheet">
<style>
:root{
  --navy:#1F3864;--blue:#2E75B6;--lblue:#D6E4F0;--teal:#0D7377;
  --orange:#E67E22;--green:#27AE60;--red:#C0392B;--purple:#6C3483;
  --bg:#EEF2F7;--card:#fff;--border:#DDE3EC;--text:#2C3E50;
  --gray:#7F8C8D;--lgray:#F5F7FA;
  --mono:'JetBrains Mono',monospace;--sans:'Noto Sans SC',sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:var(--sans);background:var(--bg);color:var(--text);min-height:100vh}

.topbar{
  background:linear-gradient(135deg,var(--navy),#2a4a82);
  padding:0 2rem;height:58px;display:flex;align-items:center;
  justify-content:space-between;position:sticky;top:0;z-index:200;
  box-shadow:0 2px 20px rgba(31,56,100,.25);
}
.brand{display:flex;align-items:center;gap:12px}
.brand-logo{width:36px;height:36px;background:var(--teal);border-radius:9px;
  display:flex;align-items:center;justify-content:center;font-size:18px}
.brand-name{font-size:15px;font-weight:700;color:#fff}
.brand-sub{font-size:11px;color:rgba(255,255,255,.55);margin-top:1px}
.topbar-right{display:flex;align-items:center;gap:10px}
.master-pill{
  display:flex;align-items:center;gap:7px;background:rgba(255,255,255,.1);
  border:1px solid rgba(255,255,255,.15);border-radius:20px;
  padding:5px 12px;font-size:12px;color:rgba(255,255,255,.85);cursor:pointer;
  transition:.2s;
}
.master-pill:hover{background:rgba(255,255,255,.18)}
.mdot{width:7px;height:7px;border-radius:50%;background:#6E7681;transition:.3s}
.mdot.ok{background:#3FB950}
.mdot.no{background:#D29922;animation:blink 1.5s infinite}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.4}}
.ver{font-family:var(--mono);font-size:11px;color:rgba(255,255,255,.45);
  background:rgba(255,255,255,.07);padding:3px 8px;border-radius:4px}

.main{max-width:980px;margin:0 auto;padding:2rem 1.5rem}
@media(max-width:640px){.main{padding:1rem}}

.card{background:var(--card);border-radius:14px;border:1px solid var(--border);
  padding:1.5rem;margin-bottom:1.25rem;box-shadow:0 1px 6px rgba(0,0,0,.05)}
.card-hdr{display:flex;align-items:center;gap:10px;margin-bottom:1.1rem;
  padding-bottom:.8rem;border-bottom:1px solid var(--border)}
.card-icon{width:34px;height:34px;border-radius:8px;display:flex;
  align-items:center;justify-content:center;font-size:17px;flex-shrink:0}
.card-title{font-size:15px;font-weight:700;color:var(--navy)}
.card-sub{font-size:12px;color:var(--gray);margin-top:2px}

/* Upload zone */
.upload-zone{border:2px dashed var(--border);border-radius:10px;padding:1.75rem;
  text-align:center;cursor:pointer;transition:.2s;background:#FAFCFF;position:relative}
.upload-zone:hover,.upload-zone.drag{border-color:var(--blue);background:#EBF3FB}
.upload-zone.has{border-color:var(--green);background:#F0FBF4;border-style:solid}
.upload-zone input[type=file]{position:absolute;inset:0;opacity:0;
  cursor:pointer;width:100%;height:100%}
.up-icon{font-size:2.2rem;margin-bottom:.5rem}
.up-label{font-size:13px;color:var(--gray)}.up-label strong{color:var(--blue)}
.up-ok{font-size:13px;color:var(--green);font-weight:600;
  display:flex;align-items:center;justify-content:center;gap:7px}

/* Form fields */
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
@media(max-width:580px){.form-grid{grid-template-columns:1fr}}
.form-grid-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px}
@media(max-width:640px){.form-grid-3{grid-template-columns:1fr}}
.field{display:flex;flex-direction:column;gap:5px}
.field label{font-size:12px;font-weight:600;color:var(--navy)}
.field input,.field select,.field textarea{
  padding:10px 13px;border:1.5px solid var(--border);border-radius:8px;
  font-family:var(--sans);font-size:13px;color:var(--text);
  background:#fff;outline:none;transition:.2s;
}
.field input:focus,.field select:focus,.field textarea:focus{
  border-color:var(--blue);box-shadow:0 0 0 3px rgba(46,117,182,.1)}
.field .hint{font-size:11px;color:var(--gray)}
.field-mono input{font-family:var(--mono);font-size:12px}

/* Buttons */
.btn{padding:10px 18px;border-radius:8px;font-family:var(--sans);font-size:13px;
  font-weight:600;cursor:pointer;border:none;transition:all .2s;
  display:inline-flex;align-items:center;gap:6px;white-space:nowrap}
.btn-blue{background:var(--blue);color:#fff}
.btn-blue:hover:not(:disabled){background:var(--navy);transform:translateY(-1px)}
.btn-blue:disabled{background:#B0C4DE;cursor:not-allowed}
.btn-teal{background:var(--teal);color:#fff}
.btn-teal:hover:not(:disabled){background:#0A5F63}
.btn-teal:disabled{background:#aaa;cursor:not-allowed}
.btn-green{background:var(--green);color:#fff}
.btn-green:hover{background:#219A52}
.btn-outline{background:#fff;color:var(--blue);border:1.5px solid var(--blue)}
.btn-outline:hover{background:var(--lblue)}
.btn-outline-red{background:#fff;color:var(--red);border:1.5px solid var(--red)}
.btn-outline-red:hover{background:#FCE4D6}
.btn-ghost{background:transparent;color:var(--gray);border:1.5px solid var(--border)}
.btn-ghost:hover{background:var(--lgray)}
.btn-sm{padding:6px 12px;font-size:12px;border-radius:6px}
.btn-lg{padding:13px 28px;font-size:15px;border-radius:10px}
.btn-block{width:100%;justify-content:center}

/* Container group UI */
.ctnr-table{width:100%;border-collapse:collapse;font-size:13px}
.ctnr-table th{background:var(--navy);color:#fff;padding:9px 12px;
  font-size:12px;font-weight:600;text-align:left}
.ctnr-table td{padding:9px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
.ctnr-table tr:last-child td{border-bottom:none}
.ctnr-table tr:hover td{background:#F8FAFF}
.group-badge{display:inline-flex;align-items:center;gap:5px;padding:3px 10px;
  border-radius:12px;font-size:11px;font-weight:700;cursor:pointer;border:none;transition:.2s}
.seq-num{width:28px;height:28px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-weight:700;font-size:13px;flex-shrink:0}

/* Group color palette */
.g0{background:#E8F4FD;color:#1A5276;border:1.5px solid #2E75B6}
.g1{background:#E9F7EF;color:#1E8449;border:1.5px solid #27AE60}
.g2{background:#FEF9E7;color:#7D6608;border:1.5px solid #D4AC0D}
.g3{background:#F5EEF8;color:#512E5F;border:1.5px solid #8E44AD}
.g4{background:#FDEBD0;color:#784212;border:1.5px solid #E67E22}
.g5{background:#FDEDEC;color:#922B21;border:1.5px solid #C0392B}
.sn0{background:#2E75B6;color:#fff}
.sn1{background:#27AE60;color:#fff}
.sn2{background:#D4AC0D;color:#fff}
.sn3{background:#8E44AD;color:#fff}
.sn4{background:#E67E22;color:#fff}
.sn5{background:#C0392B;color:#fff}

/* Progress/Result */
.prog-card{display:none}.prog-card.show{display:block}
.res-card{display:none}.res-card.show{display:block}
.terminal{background:#0D1117;border-radius:10px;padding:1rem 1.25rem;
  font-family:var(--mono);font-size:12px;line-height:1.9;
  max-height:280px;overflow-y:auto;color:#C9D1D9}
.ll{display:flex;gap:10px;align-items:baseline}
.lt{color:#484F58;flex-shrink:0;font-size:11px}
.lo{color:#3FB950}.lw{color:#D29922}.le{color:#F85149}
.li{color:#58A6FF}.ld{color:#6E7681}
.sdot{width:9px;height:9px;border-radius:50%;background:#6E7681;flex-shrink:0}
.sdot.run{background:var(--orange);animation:pulse 1s infinite}
.sdot.ok{background:var(--green)}.sdot.err{background:var(--red)}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.5;transform:scale(.8)}}
.res-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:1rem}
@media(max-width:540px){.res-grid{grid-template-columns:1fr}}
.rf{border:1.5px solid var(--border);border-radius:10px;padding:1rem;display:flex;
  flex-direction:column;gap:7px}
.rf.s1{border-color:var(--blue);background:linear-gradient(135deg,#F5FAFF,#EBF3FB)}
.rf.s2{border-color:var(--green);background:linear-gradient(135deg,#F5FDF7,#EAFAF1)}
.rf-tag{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.s1 .rf-tag{color:var(--blue)}.s2 .rf-tag{color:var(--green)}
.rf-name{font-size:12px;font-family:var(--mono);word-break:break-all;color:var(--text)}
.rf-sheets{font-size:11px;color:var(--gray)}

/* Tabs */
.tabs{display:flex;gap:0;border-bottom:2px solid var(--border);margin-bottom:1.25rem}
.tab{padding:9px 18px;font-size:13px;font-weight:600;color:var(--gray);
  cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-2px;transition:.2s}
.tab.active{color:var(--blue);border-bottom-color:var(--blue)}
.tab-panel{display:none}.tab-panel.active{display:block}

/* Modal */
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);
  z-index:300;align-items:center;justify-content:center}
.modal-bg.show{display:flex}
.modal{background:#fff;border-radius:16px;padding:2rem;width:100%;max-width:560px;
  margin:1rem;box-shadow:0 20px 60px rgba(0,0,0,.2);animation:slideUp .25s ease}
@keyframes slideUp{from{transform:translateY(16px);opacity:0}to{transform:none;opacity:1}}
.modal-title{font-size:17px;font-weight:700;color:var(--navy);margin-bottom:.35rem}
.modal-sub{font-size:13px;color:var(--gray);margin-bottom:1.4rem}
.modal-footer{display:flex;gap:10px;justify-content:flex-end;margin-top:1.5rem}
.divider{height:1px;background:var(--border);margin:1.25rem 0}

/* Tag */
.tag{display:inline-flex;align-items:center;padding:2px 8px;border-radius:10px;
  font-size:11px;font-weight:600}
.tag-blue{background:#EBF3FB;color:var(--blue)}
.tag-green{background:#E9F7EF;color:var(--green)}
.tag-orange{background:#FDEBD0;color:#784212}

/* History */
.hist-item{display:flex;align-items:center;justify-content:space-between;
  padding:10px 14px;background:var(--lgray);border-radius:8px;
  border:1px solid var(--border);gap:8px;flex-wrap:wrap;margin-bottom:8px}
.hist-batch{font-size:13px;font-weight:600;font-family:var(--mono);color:var(--navy)}
.hist-time{font-size:11px;color:var(--gray)}
</style>
</head>
<body>

<nav class="topbar">
  <div class="brand">
    <div class="brand-logo">📦</div>
    <div>
      <div class="brand-name">外贸单据自动生成系统</div>
      <div class="brand-sub">广东擎烽电气科技有限公司 · Trade Document Automation</div>
    </div>
  </div>
  <div class="topbar-right">
    <div class="master-pill" onclick="openMasterModal()">
      <div class="mdot no" id="mdot"></div>
      <span id="masterTxt">主数据未上传</span>
    </div>
    <span class="ver">v3.0</span>
  </div>
</nav>

<main class="main">

<!-- ── 装箱清单上传 ── -->
<div class="card">
  <div class="card-hdr">
    <div class="card-icon" style="background:#FEF9E7">📦</div>
    <div>
      <div class="card-title">Step 1 · 上传装箱清单</div>
      <div class="card-sub">仓管员填写后提交的当票装箱清单（.xlsx）</div>
    </div>
  </div>
  <div class="upload-zone" id="packZone">
    <input type="file" id="packFile" accept=".xlsx,.xls" onchange="handlePack(this)">
    <div class="up-icon">🗂️</div>
    <div class="up-label">拖拽文件到此处，或 <strong>点击选择</strong></div>
    <div style="font-size:11px;color:var(--gray);margin-top:4px">支持 .xlsx · .xls</div>
  </div>
</div>

<!-- ── 出货信息填写 ── -->
<div class="card" id="orderCard" style="display:none">
  <div class="card-hdr">
    <div class="card-icon" style="background:#EBF3FB">📋</div>
    <div>
      <div class="card-title">Step 2 · 填写出货信息</div>
      <div class="card-sub">选择客户和卖方，填写本票单据信息</div>
    </div>
  </div>

  <!-- Tabs -->
  <div class="tabs">
    <div class="tab active" onclick="switchTab('basic')">基础信息</div>
    <div class="tab" onclick="switchTab('containers')">货柜分组</div>
    <div class="tab" onclick="switchTab('advanced')">高级设置</div>
  </div>

  <!-- Tab: 基础信息 -->
  <div class="tab-panel active" id="tab-basic">
    <div class="form-grid" style="margin-bottom:14px">
      <div class="field">
        <label>买方客户 ★</label>
        <select id="buyerSel" onchange="updateBuyer()">
          <option value="">— 选择客户 —</option>
        </select>
        <span class="hint">选完自动带入付款条件和贸易术语</span>
      </div>
      <div class="field">
        <label>卖方主体 ★</label>
        <select id="sellerSel">
          <option value="">— 选择卖方 —</option>
        </select>
      </div>
    </div>
    <div class="form-grid" style="margin-bottom:14px">
      <div class="field">
        <label>成品编码 ★ <span style="font-size:11px;color:var(--gray)">多个产品用 | 分隔</span></label>
        <select id="productSel" onchange="checkMultiProduct()">
          <option value="">— 选择成品 —</option>
        </select>
      </div>
      <div class="field">
        <label>出货日期 ★ <span style="font-size:11px;color:var(--gray)">(Invoice Date)</span></label>
        <input type="date" id="shipDate" value="">
      </div>
    </div>
    <div class="form-grid" style="margin-bottom:14px">
      <div class="field field-mono">
        <label>Invoice No. ★</label>
        <input type="text" id="invoiceNo" placeholder="如：CF-26-US13060101A">
      </div>
      <div class="field">
        <label>报关整机套数 ★ <span style="font-size:11px;color:var(--gray)">多产品用 | 分隔</span></label>
        <input type="text" id="customsSuits" placeholder="如：740 或 440|200">
      </div>
    </div>
    <!-- 动态付款条件显示 -->
    <div id="buyerInfo" style="display:none;padding:10px 14px;background:var(--lgray);
      border-radius:8px;font-size:12px;color:var(--gray);border:1px solid var(--border);
      display:flex;gap:16px;flex-wrap:wrap">
    </div>
  </div>

  <!-- Tab: 货柜分组 -->
  <div class="tab-panel" id="tab-containers">
    <div style="margin-bottom:12px;padding:10px 14px;background:#FFF3CD;border-radius:8px;
      border-left:3px solid var(--orange);font-size:12px;color:#7D6608">
      <strong>货柜分组规则：</strong>同一颜色的柜合并出一套单据，不同颜色各自出一套。
      点击颜色标签切换分组，拖拽右侧列可排序。
    </div>
    <table class="ctnr-table" id="ctnrTable">
      <thead>
        <tr>
          <th style="width:50px">序号</th>
          <th>货柜号</th>
          <th>铅封号</th>
          <th style="width:90px">柜型</th>
          <th style="width:100px">VGM kg</th>
          <th style="width:140px">SO号</th>
          <th style="width:130px">单据分组</th>
        </tr>
      </thead>
      <tbody id="ctnrBody">
        <tr><td colspan="7" style="text-align:center;color:var(--gray);padding:2rem">
          请先上传装箱清单，货柜信息将自动读取
        </td></tr>
      </tbody>
    </table>
    <div style="margin-top:12px;display:flex;gap:8px;flex-wrap:wrap;align-items:center">
      <span style="font-size:12px;color:var(--gray)">分组颜色：</span>
      <span class="tag tag-blue">● 分组A</span>
      <span class="tag tag-green">● 分组B</span>
      <span class="tag tag-orange">● 分组C</span>
      <span style="font-size:12px;color:var(--gray);margin-left:8px">
        点击柜行的「分组」标签切换颜色
      </span>
    </div>
  </div>

  <!-- Tab: 高级设置 -->
  <div class="tab-panel" id="tab-advanced">
    <div class="form-grid-3">
      <div class="field">
        <label>付款条件</label>
        <input type="text" id="payTerms" placeholder="T/T 45 days after B/L">
      </div>
      <div class="field">
        <label>贸易术语</label>
        <select id="incoterms">
          <option>FOB</option><option>CIF</option><option>EXW</option><option>DAP</option>
        </select>
      </div>
      <div class="field">
        <label>货币</label>
        <select id="currency">
          <option>USD</option><option>CNY</option><option>EUR</option>
        </select>
      </div>
    </div>
    <div style="margin-top:14px" class="form-grid">
      <div class="field">
        <label>合同日期偏移（天，负数=早于发票）</label>
        <input type="number" id="contractOffset" value="-30">
      </div>
      <div class="field">
        <label>批次号（可自动生成）</label>
        <input type="text" id="batchNo" placeholder="留空自动生成">
      </div>
    </div>
  </div>

  <div class="divider"></div>
  <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:center">
    <button class="btn btn-blue btn-lg" id="genBtn" onclick="generate()" disabled>
      ⚡ 生成单据
    </button>
    <div id="readyHint" style="font-size:12px;color:var(--gray)">
      请完成上方必填项
    </div>
  </div>
</div>

<!-- ── 进度区 ── -->
<div class="card prog-card" id="progCard">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:1rem">
    <div class="sdot run" id="sdot"></div>
    <div>
      <div style="font-size:15px;font-weight:700;color:var(--navy)" id="progTitle">正在生成…</div>
      <div style="font-size:12px;color:var(--gray)" id="progSub">请稍候</div>
    </div>
  </div>
  <div class="terminal" id="termLog"></div>
</div>

<!-- ── 结果区 ── -->
<div class="card res-card" id="resCard">
  <div class="card-hdr">
    <span style="font-size:22px">🎉</span>
    <div>
      <div class="card-title">单据生成完成</div>
      <div class="card-sub" id="resSub"></div>
    </div>
  </div>
  <div class="res-grid" id="resGrid"></div>
</div>

<!-- ── 历史记录 ── -->
<div class="card" id="histCard" style="display:none">
  <div class="card-hdr">
    <span style="font-size:18px">🕐</span>
    <div class="card-title">本次会话记录</div>
  </div>
  <div id="histList"></div>
</div>

</main>

<!-- ═══ 主数据上传 Modal ═══ -->
<div class="modal-bg" id="masterModal">
  <div class="modal">
    <div class="modal-title">上传主数据管理手册</div>
    <div class="modal-sub">上传一次后系统长期记住。如有更新，重新上传覆盖即可。</div>
    <div class="upload-zone" id="masterZone" style="padding:1.4rem">
      <input type="file" id="masterFile" accept=".xlsx,.xls" onchange="handleMaster(this)">
      <div class="up-icon" id="mzIcon">📄</div>
      <div id="mzText">
        <div class="up-label">拖拽主数据管理手册到此<br>或 <strong>点击选择</strong></div>
        <div style="font-size:11px;color:var(--gray);margin-top:4px">主数据管理手册_最终版.xlsx</div>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-ghost" onclick="closeMasterModal()">关闭</button>
    </div>
  </div>
</div>

<!-- ═══ 新增客户 Modal ═══ -->
<div class="modal-bg" id="newBuyerModal">
  <div class="modal" style="max-width:600px">
    <div class="modal-title">新增买方客户</div>
    <div class="modal-sub">填写后自动保存到主数据手册，下次直接下拉选择。</div>
    <div class="form-grid" style="margin-bottom:12px">
      <div class="field">
        <label>客户代码 ★ <span style="font-size:11px;color:var(--gray)">如：US13</span></label>
        <input type="text" id="nb_code" placeholder="自定义简称">
      </div>
      <div class="field">
        <label>买方英文全称 ★</label>
        <input type="text" id="nb_name_en" placeholder="COMPANY NAME CO., LTD.">
      </div>
    </div>
    <div class="field" style="margin-bottom:12px">
      <label>买方地址（英文）★</label>
      <textarea id="nb_addr" rows="2" placeholder="No.XX, Street Name, City, Country" style="resize:vertical"></textarea>
    </div>
    <div class="form-grid-3" style="margin-bottom:12px">
      <div class="field">
        <label>付款条件</label>
        <input type="text" id="nb_payment" placeholder="T/T 45 days after B/L">
      </div>
      <div class="field">
        <label>贸易术语</label>
        <select id="nb_inco">
          <option>FOB</option><option>CIF</option><option>EXW</option>
        </select>
      </div>
      <div class="field">
        <label>货币</label>
        <select id="nb_currency">
          <option>USD</option><option>CNY</option><option>EUR</option>
        </select>
      </div>
    </div>
    <div class="form-grid" style="margin-bottom:12px">
      <div class="field">
        <label>联系电话</label>
        <input type="text" id="nb_tel" placeholder="+66-XX-XXXXXXXX">
      </div>
      <div class="field">
        <label>联系邮箱</label>
        <input type="text" id="nb_email" placeholder="contact@example.com">
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn btn-ghost" onclick="closeNewBuyerModal()">取消</button>
      <button class="btn btn-teal" onclick="saveNewBuyer()">💾 保存并使用</button>
    </div>
  </div>
</div>

<script>
// ── 状态 ──────────────────────────────────────────────────────────────────
const S = {
  masterLoaded: false,
  masterName: '',
  masterData: null,   // {buyers, sellers, products}
  packingPath: null,
  containers: [],     // 从装箱清单解析的货柜列表
  containerGroups: {}, // seq -> groupIndex (0,1,2...)
  history: [],
};
const GRP_NAMES = ['A','B','C','D','E','F'];
const GRP_CLS   = ['g0','g1','g2','g3','g4','g5'];
const SN_CLS    = ['sn0','sn1','sn2','sn3','sn4','sn5'];

// ── 初始化 ─────────────────────────────────────────────────────────────────
async function init() {
  document.getElementById('shipDate').value = new Date().toISOString().split('T')[0];
  try {
    const r = await fetch('/master/info');
    const d = await r.json();
    if (d.ok) {
      S.masterLoaded = true; S.masterName = d.meta.name;
      S.masterData = d.data;
      updateMasterUI(true);
      populateSelects();
    }
  } catch(e) {}
}
init();

// ── 主数据 ────────────────────────────────────────────────────────────────
function openMasterModal()  { document.getElementById('masterModal').classList.add('show'); }
function closeMasterModal() { document.getElementById('masterModal').classList.remove('show'); }
document.getElementById('masterModal').addEventListener('click', e => {
  if(e.target===e.currentTarget) closeMasterModal();
});

async function handleMaster(input) {
  const file = input.files[0]; if(!file) return;
  document.getElementById('mzIcon').textContent='⏳';
  document.getElementById('mzText').innerHTML=`<div class="up-label">正在上传并解析 ${file.name}…</div>`;
  const fd=new FormData(); fd.append('file',file);
  try {
    const r=await fetch('/master/upload',{method:'POST',body:fd});
    const d=await r.json();
    if(d.ok) {
      S.masterLoaded=true; S.masterName=d.name; S.masterData=d.data;
      updateMasterUI(true); populateSelects(); closeMasterModal();
    } else {
      document.getElementById('mzIcon').textContent='❌';
      document.getElementById('mzText').innerHTML=`<div style="color:var(--red);font-size:13px">解析失败：${d.error}</div>`;
    }
  } catch(e) {
    document.getElementById('mzIcon').textContent='❌';
    document.getElementById('mzText').innerHTML=`<div style="color:var(--red)">网络错误</div>`;
  }
}

function updateMasterUI(loaded) {
  const dot=document.getElementById('mdot');
  const txt=document.getElementById('masterTxt');
  if(loaded) {
    dot.className='mdot ok'; txt.textContent=S.masterName||'主数据已就绪';
  } else {
    dot.className='mdot no'; txt.textContent='主数据未上传';
  }
  updateGenBtn();
}

function populateSelects() {
  if(!S.masterData) return;
  const bs=document.getElementById('buyerSel');
  const ss=document.getElementById('sellerSel');
  const ps=document.getElementById('productSel');

  // 保留现有值
  const prevB=bs.value, prevS=ss.value, prevP=ps.value;

  bs.innerHTML='<option value="">— 选择客户 —</option>';
  (S.masterData.buyers||[]).forEach(b=>{
    bs.innerHTML+=`<option value="${b.code}" data-payment="${b.payment_terms}"
      data-inco="${b.incoterms}" data-currency="${b.currency}"
      data-addr="${esc(b.address_en)}">${b.code} · ${b.name_en.substring(0,30)}</option>`;
  });
  bs.innerHTML+=`<option value="__new__">＋ 新增买方客户…</option>`;
  if(prevB) bs.value=prevB;

  ss.innerHTML='<option value="">— 选择卖方 —</option>';
  (S.masterData.sellers||[]).forEach(s=>{
    ss.innerHTML+=`<option value="${s.code}">${s.code} · ${s.name_cn}</option>`;
  });
  if(prevS) ss.value=prevS;
  else if((S.masterData.sellers||[]).length===1) ss.value=S.masterData.sellers[0].code;

  ps.innerHTML='<option value="">— 选择成品 —</option>';
  (S.masterData.products||[]).forEach(p=>{
    ps.innerHTML+=`<option value="${p.code}">${p.code} · ${p.name_cn}</option>`;
  });
  if(prevP) ps.value=prevP;
}

function updateBuyer() {
  const sel=document.getElementById('buyerSel');
  if(sel.value==='__new__') {
    sel.value=''; openNewBuyerModal(); return;
  }
  const opt=sel.selectedOptions[0];
  if(!opt||!opt.value) { document.getElementById('buyerInfo').style.display='none'; return; }
  document.getElementById('payTerms').value=opt.dataset.payment||'';
  document.getElementById('incoterms').value=opt.dataset.inco||'FOB';
  document.getElementById('currency').value=opt.dataset.currency||'USD';
  const info=document.getElementById('buyerInfo');
  info.style.display='flex';
  info.innerHTML=`<span>💳 ${opt.dataset.payment||'—'}</span>
    <span>🚢 ${opt.dataset.inco||'FOB'}</span>
    <span>💵 ${opt.dataset.currency||'USD'}</span>`;
  updateGenBtn();
}

// ── 装箱清单上传 ──────────────────────────────────────────────────────────
async function handlePack(input) {
  const file=input.files[0]; if(!file) return;
  const zone=document.getElementById('packZone');
  zone.innerHTML=`<div style="font-size:13px;color:var(--gray)">⏳ 正在上传并解析 ${file.name}…</div>
    <input type="file" accept=".xlsx,.xls" onchange="handlePack(this)"
      style="position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%">`;
  const fd=new FormData(); fd.append('file',file); fd.append('type','packing');
  try {
    const r=await fetch('/upload',{method:'POST',body:fd});
    const d=await r.json();
    if(d.ok) {
      S.packingPath=d.path;
      zone.className='upload-zone has';
      zone.innerHTML=`<div class="up-ok">✅ ${file.name}</div>
        <div style="font-size:11px;color:var(--gray);margin-top:4px">${d.size} · ${d.container_count} 个货柜 · ${d.row_count} 行明细</div>
        <input type="file" accept=".xlsx,.xls" onchange="handlePack(this)"
          style="position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%">`;
      S.containers=d.containers||[];
      renderContainerTable();
      document.getElementById('orderCard').style.display='block';
      document.getElementById('orderCard').scrollIntoView({behavior:'smooth',block:'nearest'});
    } else {
      zone.innerHTML=`<div style="color:var(--red);font-size:13px">❌ ${d.error}</div>
        <input type="file" accept=".xlsx,.xls" onchange="handlePack(this)"
          style="position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%">`;
    }
  } catch(e) {
    zone.innerHTML=`<div style="color:var(--red);font-size:13px">❌ 网络错误</div>
      <input type="file" accept=".xlsx,.xls" onchange="handlePack(this)"
        style="position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%">`;
  }
  updateGenBtn();
}

// ── 货柜分组表格 ──────────────────────────────────────────────────────────
function renderContainerTable() {
  const tbody=document.getElementById('ctnrBody');
  if(!S.containers.length) {
    tbody.innerHTML='<tr><td colspan="7" style="text-align:center;color:var(--gray);padding:2rem">未解析到货柜信息</td></tr>';
    return;
  }
  // 初始化分组（每柜默认各自一组）
  S.containers.forEach((c,i)=>{
    if(S.containerGroups[c.seq]===undefined) S.containerGroups[c.seq]=i;
  });
  tbody.innerHTML=S.containers.map(c=>{
    const gi=S.containerGroups[c.seq]||0;
    const gc=GRP_CLS[gi%GRP_CLS.length];
    const snc=SN_CLS[gi%SN_CLS.length];
    const gname=GRP_NAMES[gi%GRP_NAMES.length];
    return `<tr>
      <td><div class="seq-num ${snc}">${c.seq}</div></td>
      <td><span style="font-family:var(--mono);font-size:12px">${c.container_no||'—'}</span></td>
      <td style="font-size:12px;color:var(--gray)">${c.seal_no||'—'}</td>
      <td><span class="tag tag-blue">${c.container_size||'40HQ'}</span></td>
      <td style="font-size:12px">${c.vgm_kg||'—'}</td>
      <td style="font-family:var(--mono);font-size:11px;color:var(--gray)">${c.so_no||'—'}</td>
      <td>
        <button class="group-badge ${gc}" onclick="cycleGroup(${c.seq})"
          id="gbtn-${c.seq}">
          ● 分组${gname}
        </button>
      </td>
    </tr>`;
  }).join('');
}

function cycleGroup(seq) {
  const curGi=S.containerGroups[seq]||0;
  // 找当前最大分组+1，或循环
  const maxG=Math.max(...Object.values(S.containerGroups))||0;
  const nextG=(curGi===maxG && maxG<5) ? maxG+1 : (curGi+1)%(maxG+1===1?2:maxG+2);
  // 简单逻辑：点击就+1，超过当前分组数则重置到0
  const allGs=new Set(Object.values(S.containerGroups));
  let ng=curGi+1;
  if(ng>allGs.size) ng=0;
  S.containerGroups[seq]=ng;
  renderContainerTable();
}

// ── 页签切换 ──────────────────────────────────────────────────────────────
function switchTab(name) {
  document.querySelectorAll('.tab').forEach((t,i)=>{
    const names=['basic','containers','advanced'];
    t.className='tab'+(names[i]===name?' active':'');
  });
  document.querySelectorAll('.tab-panel').forEach(p=>p.className='tab-panel');
  document.getElementById('tab-'+name).className='tab-panel active';
}

// ── 新增买方 ──────────────────────────────────────────────────────────────
function openNewBuyerModal()  { document.getElementById('newBuyerModal').classList.add('show'); }
function closeNewBuyerModal() { document.getElementById('newBuyerModal').classList.remove('show'); }
document.getElementById('newBuyerModal').addEventListener('click', e=>{
  if(e.target===e.currentTarget) closeNewBuyerModal();
});

async function saveNewBuyer() {
  const data={
    code:  document.getElementById('nb_code').value.trim(),
    name_en: document.getElementById('nb_name_en').value.trim(),
    address_en: document.getElementById('nb_addr').value.trim(),
    payment_terms: document.getElementById('nb_payment').value.trim(),
    incoterms: document.getElementById('nb_inco').value,
    currency:  document.getElementById('nb_currency').value,
    tel:   document.getElementById('nb_tel').value.trim(),
    email: document.getElementById('nb_email').value.trim(),
  };
  if(!data.code||!data.name_en||!data.address_en) {
    alert('客户代码、英文全称、地址为必填项'); return;
  }
  try {
    const r=await fetch('/buyer/add',{method:'POST',
      headers:{'Content-Type':'application/json'},body:JSON.stringify(data)});
    const d=await r.json();
    if(d.ok) {
      if(d.data) { S.masterData=d.data; populateSelects(); }
      document.getElementById('buyerSel').value=data.code;
      updateBuyer();
      closeNewBuyerModal();
      // 清空表单
      ['nb_code','nb_name_en','nb_addr','nb_payment','nb_tel','nb_email'].forEach(id=>{
        document.getElementById(id).value='';
      });
    } else { alert('保存失败：'+d.error); }
  } catch(e) { alert('网络错误'); }
}

// ── 生成按钮状态 ──────────────────────────────────────────────────────────
function updateGenBtn() {
  const ready = S.masterLoaded && S.packingPath &&
    document.getElementById('buyerSel').value &&
    document.getElementById('sellerSel').value &&
    document.getElementById('productSel').value &&
    document.getElementById('invoiceNo').value.trim() &&
    document.getElementById('customsSuits').value.trim();
  document.getElementById('genBtn').disabled=!ready;
  document.getElementById('readyHint').textContent=
    ready ? '准备就绪，点击生成' : '请完成上方必填项（★）';
  document.getElementById('readyHint').style.color=ready?'var(--green)':'var(--gray)';
}

// 监听必填字段
['buyerSel','sellerSel','productSel','invoiceNo','customsSuits'].forEach(id=>{
  const el=document.getElementById(id);
  if(el) el.addEventListener('input',updateGenBtn), el.addEventListener('change',updateGenBtn);
});

// ── 生成 ──────────────────────────────────────────────────────────────────
async function generate() {
  // 构造分组信息：{groupIndex: [seq1, seq2, ...]}
  const groups={};
  Object.entries(S.containerGroups).forEach(([seq,gi])=>{
    if(!groups[gi]) groups[gi]=[];
    groups[gi].push(parseInt(seq));
  });

  // 构造批次列表（每个分组对应一套单据）
  const batchBase=document.getElementById('batchNo').value.trim() ||
    `${document.getElementById('buyerSel').value}-${document.getElementById('shipDate').value.replace(/-/g,'').slice(2)}`;

  const batches=Object.entries(groups).map(([gi,seqs],i)=>({
    batch_no:     `${batchBase}-${String.fromCharCode(65+parseInt(gi))}`,
    customer_code:document.getElementById('buyerSel').value,
    shipment_date:document.getElementById('shipDate').value,
    invoice_no:   document.getElementById('invoiceNo').value.trim() + (Object.keys(groups).length>1?String.fromCharCode(65+parseInt(gi)):''),
    product_codes:document.getElementById('productSel').value.split('|').map(s=>s.trim()),
    customs_suits:document.getElementById('customsSuits').value.split('|').map(s=>parseInt(s.trim())||0),
    container_seqs:seqs,
    seller_code:  document.getElementById('sellerSel').value,
    payment_terms:document.getElementById('payTerms').value,
    incoterms:    document.getElementById('incoterms').value,
    contract_offset:parseInt(document.getElementById('contractOffset').value)||0,
  }));

  // 重置UI
  const prog=document.getElementById('progCard');
  const res=document.getElementById('resCard');
  prog.classList.add('show'); res.classList.remove('show');
  document.getElementById('termLog').innerHTML='';
  document.getElementById('sdot').className='sdot run';
  document.getElementById('progTitle').textContent='正在生成单据…';
  document.getElementById('progSub').textContent=`${batches.length} 套单据`;
  document.getElementById('genBtn').disabled=true;
  prog.scrollIntoView({behavior:'smooth',block:'nearest'});

  try {
    const r=await fetch('/generate_v3',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({packing_path:S.packingPath,batches})
    });
    const d=await r.json();
    if(!d.ok){addLog('e',d.error);document.getElementById('sdot').className='sdot err';document.getElementById('genBtn').disabled=false;return;}
    pollTask(d.task_id);
  } catch(e) {
    addLog('e','请求失败：'+e.message);
    document.getElementById('sdot').className='sdot err';
    document.getElementById('genBtn').disabled=false;
  }
}

async function pollTask(tid) {
  const poll=async()=>{
    try {
      const r=await fetch(`/task/${tid}`);
      const d=await r.json();
      (d.logs||[]).forEach(l=>{
        const c=l.includes('✅')?'o':l.includes('⚠')?'w':l.includes('✗')?'e':
                 (l.includes('📦')||l.includes('📄')||l.includes('🔗'))?'i':'d';
        addLog(c,l);
      });
      if(d.status==='running'){setTimeout(poll,800);return;}
      if(d.status==='done'){
        document.getElementById('sdot').className='sdot ok';
        document.getElementById('progTitle').textContent='✅ 生成完成';
        document.getElementById('progSub').textContent=`${d.outputs.length} 套单据已生成`;
        showResults(d.outputs);
      } else {
        document.getElementById('sdot').className='sdot err';
        document.getElementById('progTitle').textContent='❌ 生成失败';
      }
      document.getElementById('genBtn').disabled=false;
    } catch(e){setTimeout(poll,2000);}
  };
  poll();
}

function showResults(outputs) {
  const rc=document.getElementById('resCard');
  const rg=document.getElementById('resGrid');
  rc.classList.add('show');
  document.getElementById('resSub').textContent=
    `${outputs.length} 套单据 · ${new Date().toLocaleTimeString()}`;
  rg.innerHTML='';
  outputs.forEach(out=>{
    if(out.set1) rg.innerHTML+=fCard('s1',out.set1,out.set1_name,'套一（中国出口）','Inv. · PL · 报关单 · 合同 · SI');
    if(out.set2) rg.innerHTML+=fCard('s2',out.set2,out.set2_name,'套二（泰国清关）','Inv.(Parts) · PL(Parts)含图片');
  });
  S.history.unshift({label:outputs.map(o=>o.batch).join('+'),time:new Date().toLocaleTimeString(),outputs});
  renderHist();
  rc.scrollIntoView({behavior:'smooth',block:'nearest'});
}

function fCard(cls,fid,fname,tag,sheets){
  return `<div class="rf ${cls}">
    <div class="rf-tag">${tag}</div>
    <div class="rf-name">📄 ${fname}</div>
    <div class="rf-sheets">${sheets}</div>
    <button class="btn ${cls==='s1'?'btn-outline':'btn-green'} btn-block"
      onclick="dl('${fid}','${fname}')" style="margin-top:6px">⬇ 下载</button>
  </div>`;
}
async function dl(fid,fname){
  const a=document.createElement('a');a.href=`/download/${fid}`;a.download=fname;
  document.body.appendChild(a);a.click();document.body.removeChild(a);
}
function renderHist(){
  if(!S.history.length)return;
  document.getElementById('histCard').style.display='block';
  document.getElementById('histList').innerHTML=S.history.slice(0,8).map(h=>`
    <div class="hist-item">
      <div><div class="hist-batch">${h.label}</div><div class="hist-time">${h.time}</div></div>
      <div style="display:flex;gap:6px">
        ${h.outputs.flatMap(o=>[
          o.set1?`<button class="btn btn-outline btn-sm" onclick="dl('${o.set1}','${o.set1_name}')">套一</button>`:'',
          o.set2?`<button class="btn btn-green btn-sm" onclick="dl('${o.set2}','${o.set2_name}')">套二</button>`:'',
        ]).join('')}
      </div>
    </div>`).join('');
}

function addLog(cls,text){
  const log=document.getElementById('termLog');
  const t=new Date().toLocaleTimeString('zh-CN',{hour12:false});
  const d=document.createElement('div');d.className='ll';
  d.innerHTML=`<span class="lt">${t}</span><span class="l${cls}">${esc(text)}</span>`;
  log.appendChild(d);log.scrollTop=log.scrollHeight;
}
function esc(t){return t.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}

// 拖拽支持
const pz=document.getElementById('packZone');
pz.addEventListener('dragover',e=>{e.preventDefault();pz.classList.add('drag')});
pz.addEventListener('dragleave',()=>pz.classList.remove('drag'));
pz.addEventListener('drop',e=>{
  e.preventDefault();pz.classList.remove('drag');
  const f=e.dataTransfer.files[0];
  if(f){const i=pz.querySelector('input');const dt=new DataTransfer();dt.items.add(f);i.files=dt.files;handlePack(i);}
});
</script>
</body>
</html>"""


# ── API ───────────────────────────────────────────────────────────────────────

@app.route('/')
def index(): return render_template_string(HTML)


def _parse_master_data(path: str) -> dict:
    """解析主数据手册，提取买方/卖方/成品下拉数据"""
    import pandas as pd
    buyers, sellers, products = [], [], []
    try:
        xl = pd.ExcelFile(path)
        for sh in xl.sheet_names:
            # 往来方
            if '往来方' in sh:
                df = pd.read_excel(xl, sheet_name=sh, header=None)
                for _, row in df.iterrows():
                    code  = str(row[0]).strip() if pd.notna(row[0]) else ''
                    ptype = str(row[1]).strip() if pd.notna(row[1]) else ''
                    if not code or ptype not in ('SELLER','BUYER','FORWARDER'): continue
                    name_en   = str(row[2]).strip() if pd.notna(row[2]) else ''
                    name_cn   = str(row[3]).strip() if pd.notna(row[3]) else ''
                    addr_en   = str(row[4]).strip() if pd.notna(row[4]) else ''
                    payment   = str(row[8]).strip() if pd.notna(row[8]) else ''
                    incoterms = str(row[9]).strip() if pd.notna(row[9]) else 'FOB'
                    currency  = str(row[10]).strip() if pd.notna(row[10]) else 'USD'
                    tax_id    = str(row[11]).strip() if pd.notna(row[11]) else ''
                    port      = str(row[12]).strip() if pd.notna(row[12]) else ''
                    if '（' in name_en or name_en in ('nan',''):  continue
                    if ptype == 'BUYER':
                        buyers.append({'code':code,'name_en':name_en,'name_cn':name_cn,
                            'address_en':addr_en,'payment_terms':payment,
                            'incoterms':incoterms,'currency':currency})
                    elif ptype == 'SELLER':
                        sellers.append({'code':code,'name_en':name_en,'name_cn':name_cn,
                            'tax_id':tax_id,'port_loading':port})
            # 成品档案
            elif '成品档案' in sh:
                df = pd.read_excel(xl, sheet_name=sh, header=None)
                for _, row in df.iterrows():
                    code = str(row[0]).strip() if pd.notna(row[0]) else ''
                    name_cn = str(row[1]).strip() if pd.notna(row[1]) else ''
                    if not code or '成品编码' in code or '★' in code or '待填' in code: continue
                    if '（' in code: continue
                    products.append({'code':code,'name_cn':name_cn})
    except Exception as e:
        pass
    return {'buyers': buyers, 'sellers': sellers, 'products': products}


@app.route('/master/info')
def master_info():
    meta = load_master_meta()
    if meta and os.path.exists(meta.get('path','')):
        data = _parse_master_data(meta['path'])
        return jsonify(ok=True, meta=meta, data=data)
    return jsonify(ok=False)


@app.route('/master/upload', methods=['POST'])
def master_upload():
    f = request.files.get('file')
    if not f: return jsonify(ok=False, error='没有文件')
    fname = 'master_' + uuid.uuid4().hex[:8] + os.path.splitext(f.filename)[1]
    path = os.path.join(MASTER_DIR, fname)
    f.save(path)
    size = os.path.getsize(path)
    size_str = f'{size/1024:.1f} KB' if size < 1048576 else f'{size/1048576:.1f} MB'
    meta = {'path':path,'name':f.filename,'size':size_str,
            'uploaded_at':datetime.now().strftime('%Y-%m-%d %H:%M')}
    save_master_meta(meta)
    data = _parse_master_data(path)
    return jsonify(ok=True, **meta, data=data)


@app.route('/master/clear', methods=['POST'])
def master_clear():
    if os.path.exists(MASTER_META): os.remove(MASTER_META)
    return jsonify(ok=True)


@app.route('/buyer/add', methods=['POST'])
def buyer_add():
    """新增买方客户，追加到主数据手册④往来方档案"""
    data = request.json
    master_path = get_master_path()
    if not master_path:
        return jsonify(ok=False, error='请先上传主数据手册')
    try:
        from openpyxl import load_workbook
        wb = load_workbook(master_path)
        ws = None
        for name in wb.sheetnames:
            if '往来方' in name: ws = wb[name]; break
        if not ws: return jsonify(ok=False, error='未找到往来方档案Sheet')
        # 找第一个空行
        max_r = ws.max_row + 1
        row_data = [
            data['code'], 'BUYER', data['name_en'], data.get('name_cn',''),
            data['address_en'], '', '', f"Tel:{data.get('tel','')} Email:{data.get('email','')}",
            data.get('payment_terms',''), data.get('incoterms','FOB'),
            data.get('currency','USD'), '', '',
        ]
        for ci, v in enumerate(row_data, 1):
            ws.cell(row=max_r, column=ci, value=v)
        wb.save(master_path)
        new_data = _parse_master_data(master_path)
        return jsonify(ok=True, data=new_data)
    except Exception as e:
        return jsonify(ok=False, error=str(e))


@app.route('/upload', methods=['POST'])
def upload():
    f = request.files.get('file')
    if not f: return jsonify(ok=False, error='没有文件')
    uid = uuid.uuid4().hex[:8]
    ext = os.path.splitext(f.filename)[1]
    path = os.path.join(UPLOAD_DIR, f'packing_{uid}{ext}')
    f.save(path)
    size = os.path.getsize(path)
    size_str = f'{size/1024:.1f} KB' if size < 1048576 else f'{size/1048576:.1f} MB'
    # 解析货柜信息
    containers = []
    row_count = 0
    try:
        import pandas as pd
        df = pd.read_excel(path, sheet_name='装箱清单', header=None)
        for ri in range(7, 15):
            if ri >= len(df): break
            row = df.iloc[ri]
            seq = int(float(row[0])) if str(row[0]).strip() not in ('nan','') else None
            cno = str(row[1]).strip() if pd.notna(row[1]) else ''
            if not seq or not cno: continue
            containers.append({
                'seq': seq, 'container_no': cno,
                'seal_no': str(row[4]).strip() if pd.notna(row[4]) else '',
                'container_size': str(row[6]).strip() if pd.notna(row[6]) else '40HQ',
                'vgm_kg': float(row[9]) if pd.notna(row[9]) and str(row[9]).strip() not in ('nan','') else None,
                'so_no': str(row[10]).strip() if pd.notna(row[10]) else '',
                'etd': str(row[13]).strip() if pd.notna(row[13]) else '',
                'port_loading': str(row[15]).strip() if pd.notna(row[15]) else '',
            })
        # 数明细行
        for ri in range(19, len(df)):
            row = df.iloc[ri]
            if str(row[1]).strip() not in ('nan','') and str(row[2]).strip() not in ('nan',''): row_count += 1
    except Exception as e:
        pass
    return jsonify(ok=True, path=path, size=size_str,
                   containers=containers, container_count=len(containers), row_count=row_count)


@app.route('/generate_v3', methods=['POST'])
def generate_v3():
    data = request.json
    packing_path = data.get('packing_path')
    batches = data.get('batches', [])
    master_path = get_master_path()
    if not master_path: return jsonify(ok=False, error='请先上传主数据管理手册')
    if not packing_path or not os.path.exists(packing_path):
        return jsonify(ok=False, error='请先上传装箱清单')
    task_id = uuid.uuid4().hex[:12]
    tasks[task_id] = {'status':'running','logs':[],'outputs':[],'_cur':0}
    threading.Thread(target=run_task_v3,
        args=(task_id, master_path, packing_path, batches), daemon=True).start()
    return jsonify(ok=True, task_id=task_id)


@app.route('/task/<tid>')
def task_status(tid):
    t = tasks.get(tid)
    if not t: return jsonify(status='not_found')
    cur = t['_cur']
    new_logs = t['logs'][cur:]
    t['_cur'] = len(t['logs'])
    return jsonify(status=t['status'], logs=new_logs, outputs=t['outputs'])


@app.route('/download/<path:fname>')
def download_file(fname):
    safe = os.path.basename(fname)
    path = os.path.join(OUTPUT_DIR, safe)
    if not os.path.exists(path): return '文件不存在', 404
    return send_file(path, as_attachment=True, download_name=safe)


def run_task_v3(tid, master_path, packing_path, batches):
    task = tasks[tid]
    def log(msg): task['logs'].append(msg)
    try:
        log('='*50); log('  外贸单据自动化系统 v3.0')
        log(f'  {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'); log('='*50)
        log('  📋 加载主数据管理手册…')
        from engine.config import MasterData, ShipmentMain, ShipmentContainer
        master = MasterData(master_path)
        log(f'  ✅ 成品档案：{len(master.finished_goods)} 条')
        log(f'  ✅ 物料价格：{len(master.component_prices)} 条')
        log(f'  ✅ 往来方：{len(master.parties)} 个')

        from engine.parser import parse_packing_list, build_document_bundle
        from engine.output import generate_document_set
        log('  📦 解析装箱清单…')
        packing = parse_packing_list(packing_path)
        log(f'     货柜：{len(packing.containers)} 个  明细：{len(packing.rows)} 行')

        outputs = []
        for bi, batch_info in enumerate(batches):
            bn = batch_info['batch_no']
            log(f'\n  ─── 批次 {bi+1}/{len(batches)}：{bn}')

            # 动态注入 ShipmentMain/Container（来自网页表单，不从Excel读）
            seqs = batch_info.get('container_seqs', [])
            master.shipment_mains[bn] = ShipmentMain(
                batch_no=bn,
                customer_code=batch_info.get('customer_code',''),
                shipment_date=batch_info.get('shipment_date',''),
                invoice_no=batch_info.get('invoice_no',''),
                product_codes=batch_info.get('product_codes',[]),
                customs_suits=batch_info.get('customs_suits',[]),
                container_seq_str=','.join(str(s) for s in seqs),
                seller_code=batch_info.get('seller_code','QF-CN'),
                remark='',
            )
            # 货柜信息从装箱清单读
            for seq in seqs:
                if seq in packing.containers:
                    pc = packing.containers[seq]
                    master.shipment_containers[seq] = ShipmentContainer(
                        seq=seq, batch_no=bn,
                        container_no=pc.container_no,
                        seal_no=pc.seal_no,
                        container_size=pc.container_size,
                        vgm_kg=pc.vgm_kg,
                        so_no=pc.so_no,
                        etd=pc.etd,
                        port_loading=pc.port_loading,
                        remark='',
                    )

            try:
                bundle = build_document_bundle(master, bn, packing)
                log(f'  ✅ Invoice: {bundle.invoice_no}')
                log(f'     货柜：{[c.container_no for c in bundle.containers]}')
                log(f'     套一：USD {bundle.set1_total_amount:,.2f}')
                log(f'     套二：USD {bundle.set2_total_amount:,.2f}  ({len(bundle.set2_lines)} 部件)')
                log('  📄 生成 Excel…')
                paths = generate_document_set(bundle, OUTPUT_DIR, master_path=master_path)
                s1=os.path.basename(paths['set1']); s2=os.path.basename(paths['set2'])
                log(f'  ✅ 套一：{s1}'); log(f'  ✅ 套二：{s2}')
                outputs.append({'batch':bn,'set1':s1,'set1_name':s1,'set2':s2,'set2_name':s2})
            except Exception as e:
                log(f'  ✗ 异常：{e}'); log(traceback.format_exc()[:400])

        log(f'\n{"="*50}'); log(f'  ✅ 完成！{len(outputs)} 套'); log('='*50)
        task['outputs']=outputs
        task['status']='done' if outputs else 'error'
    except Exception as e:
        task['logs'].append(f'  ✗ 系统错误：{e}')
        task['logs'].append(traceback.format_exc()[:500])
        task['status']='error'


if __name__ == '__main__':
    print('\n'+'='*56)
    print('  外贸单据自动生成系统 v3.0')
    print('  网页端直接选客户/卖方/货柜分组，无需手填批次记录')
    print('='*56)
    print('  本地访问：http://localhost:5000')
    print('  按 Ctrl+C 停止')
    print('='*56+'\n')
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
