"""
output.py — 生成Excel单据（套一 + 套二 + SI），格式与现有单据一致
"""
import os
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .parser import DocumentBundle, ComponentLineItem, FinishedGoodLineItem, ContainerInfo

try:
    from num2words import num2words
    def amount_words(amount: float) -> str:
        dollars = int(amount)
        cents = round((amount - dollars) * 100)
        w = num2words(dollars, lang='en').upper()
        if cents:
            return f"TOTAL U.S. DOLLARS {w} AND CENTS {num2words(cents, lang='en').upper()} ONLY."
        return f"TOTAL U.S. DOLLARS {w} ONLY."
except ImportError:
    def amount_words(amount: float) -> str:
        return f"TOTAL U.S. DOLLARS {amount:.2f} ONLY."


# ── 样式工具 ──────────────────────────────────────────────────────────────────

def _fl(c): return PatternFill('solid', start_color=c, fgColor=c)
def _fn(bold=False, size=10, color='000000', name='Arial', italic=False):
    return Font(bold=bold, size=size, color=color, name=name, italic=italic)
def _al(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bd(c='000000', style='thin'):
    s = Side(border_style=style, color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _sc(ws, r, c, val='', bg=None, bold=False, size=10, h='left', color='000000',
        italic=False, num_fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _fn(bold, size, color, italic=italic)
    if bg: cell.fill = _fl(bg)
    cell.alignment = _al(h, 'center', True)
    cell.border = _bd()
    if num_fmt: cell.number_format = num_fmt
    return cell

def _mc(ws, r1, c1, r2, c2, val='', bg=None, bold=False, size=10,
        h='left', color='000000', italic=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font = _fn(bold, size, color, italic=italic)
    if bg: cell.fill = _fl(bg)
    cell.alignment = _al(h, 'center', True)
    return cell


# ── Sheet: Inv.（套一整机发票）────────────────────────────────────────────────

def _build_set1_invoice(ws, bundle: DocumentBundle):
    """套一发票 - 整机归并"""
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 4
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16

    r = 1
    _mc(ws,r,1,r,9, bundle.seller_name_en, bold=True, size=12)
    ws.row_dimensions[r].height = 22; r += 1
    ws.row_dimensions[r].height = 6;  r += 1
    _mc(ws,r,1,r,9, bundle.seller_address_en, size=9)
    ws.row_dimensions[r].height = 16; r += 1
    for _ in range(3): ws.row_dimensions[r].height = 6; r += 1

    _mc(ws,r,1,r,1, 'Messrs：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_name_en, bold=True)
    _sc(ws,r,7, 'Invoice No.：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_no, bold=True)
    ws.row_dimensions[r].height = 18; r += 1

    _mc(ws,r,1,r,1, 'Address：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_address_en)
    _sc(ws,r,7, 'Invoice Date：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_date)
    ws.row_dimensions[r].height = 26; r += 1

    _sc(ws,r,7, 'Payment Terms：', bold=True, h='right')
    _sc(ws,r,8, bundle.buyer_payment_terms)
    ws.row_dimensions[r].height = 16; r += 1

    _sc(ws,r,7, 'Trade Terms：', bold=True, h='right')
    _sc(ws,r,8, bundle.buyer_incoterms)
    ws.row_dimensions[r].height = 16; r += 1

    # 表头
    for ci, (h, bg) in enumerate([
        ('Mark & No.', 'D6E4F0'), ('Description of Goods', 'D6E4F0'),
        ('', 'D6E4F0'), ('', 'D6E4F0'), ('', 'D6E4F0'),
        ('Quantity (PCS)', 'D6E4F0'), ('', 'D6E4F0'),
        ('Unit Price (USD)', 'D6E4F0'), ('Amount\n(USD)', 'D6E4F0'),
    ], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _fn(True, 9); c.fill = _fl(bg)
        c.alignment = _al('center'); c.border = _bd()
    ws.row_dimensions[r].height = 28; r += 1

    # EF Mark + 货柜号
    ctnr_nos = ' / '.join(c.container_no for c in bundle.containers)
    _sc(ws,r,1, 'EF', bold=True, h='center')
    ws.row_dimensions[r].height = 16; r += 1
    _sc(ws,r,1, ctnr_nos, bold=True, h='center')
    ws.row_dimensions[r].height = 16

    # 数据行
    for line in bundle.set1_lines:
        _mc(ws,r,2,r,5, line.name_en)
        _sc(ws,r,6, int(line.customs_suits), h='center'); ws.cell(row=r,column=6).border = _bd()
        _sc(ws,r,8, line.unit_price, h='center', num_fmt='#,##0.00'); ws.cell(row=r,column=8).border = _bd()
        _sc(ws,r,9, line.total_amount, h='center', num_fmt='#,##0.00'); ws.cell(row=r,column=9).border = _bd()
        ws.row_dimensions[r].height = 18; r += 1

    r += 1
    _mc(ws,r,2,r,9, amount_words(bundle.set1_total_amount), italic=True, size=9)
    ws.row_dimensions[r].height = 16; r += 2

    _mc(ws,r,6,r,7, 'TOTAL AMOUNT:', bold=True, h='right')
    _sc(ws,r,8, 'USD', bold=True, h='center')
    _sc(ws,r,9, bundle.set1_total_amount, bold=True, h='center', num_fmt='#,##0.00')
    ws.row_dimensions[r].height = 20; r += 2

    _mc(ws,r,1,r,4, f"Signed by: {bundle.seller_name_en}", size=9)
    _mc(ws,r,6,r,9, f"Port of Loading: {bundle.seller_port_loading}", size=9)


# ── Sheet: PL（套一整机装箱单）───────────────────────────────────────────────

def _build_set1_pl(ws, bundle: DocumentBundle) -> dict:
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14

    r = 1
    _mc(ws,r,1,r,10, bundle.seller_name_en, bold=True, size=12)
    ws.row_dimensions[r].height = 22; r += 1
    ws.row_dimensions[r].height = 6; r += 1
    _mc(ws,r,1,r,10, bundle.seller_address_en, size=9)
    ws.row_dimensions[r].height = 16; r += 1
    for _ in range(3): ws.row_dimensions[r].height = 6; r += 1

    _mc(ws,r,1,r,1, 'Messrs：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_name_en, bold=True)
    _sc(ws,r,7, 'Invoice No.：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_no, bold=True)
    ws.row_dimensions[r].height = 18; r += 1
    _mc(ws,r,1,r,1, 'Address：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_address_en)
    _sc(ws,r,7, 'Invoice Date：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_date)
    ws.row_dimensions[r].height = 26; r += 1

    _mc(ws,r,3,r,10, 'PACKING  LIST', bold=True, size=14, h='center')
    ws.row_dimensions[r].height = 22; r += 1

    pl_hdrs = ['Mark & Container No.','Description of Goods','','Pituctures of Goods','',
               'Quantity \n(PCS)','Quantity \n(PKGS)','N.W.\n(KGS)','G.W.\n(KGS)','Measurement  (CBM)']
    for ci, h in enumerate(pl_hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _fn(True, 9); c.fill = _fl('D6E4F0')
        c.alignment = _al('center'); c.border = _bd()
    ws.row_dimensions[r].height = 36; r += 1

    ctnr_nos = ' / '.join(c.container_no for c in bundle.containers)
    _sc(ws,r,1, 'EF', bold=True, h='center')
    ws.row_dimensions[r].height = 16; r += 1
    _sc(ws,r,1, ctnr_nos, bold=True, h='center')
    ws.row_dimensions[r].height = 18

    first_row = r
    t_pcs=t_pkgs=t_nw=t_gw=0

    for line in bundle.set1_lines:
        _mc(ws,r,2,r,3, line.name_en)
        _sc(ws,r,6, int(line.customs_suits), h='center'); ws.cell(row=r,column=6).border = _bd()
        ws.row_dimensions[r].height = 18
        t_pcs += line.customs_suits; r += 1

    # CBM写在第一数据行
    ws.cell(row=first_row, column=10).value = bundle.total_cbm
    ws.cell(row=first_row, column=10).font = _fn(size=10)
    ws.cell(row=first_row, column=10).alignment = _al('center')

    r += 1
    from num2words import num2words as n2w
    pkgs_word = n2w(bundle.total_pkgs, lang='en').upper()
    _mc(ws,r,2,r,10,
        f"TOTAL PACKED IN {pkgs_word} ({bundle.total_pkgs}) PACKAGES ONLY.",
        italic=True, size=9)
    ws.row_dimensions[r].height = 16; r += 2

    _mc(ws,r,4,r,4, 'TOTAL:', bold=True, h='right')
    for ci, val in [(6,t_pcs),(7,bundle.total_pkgs),(8,bundle.total_nw),
                    (9,bundle.total_gw),(10,bundle.total_cbm)]:
        _sc(ws,r,ci, val, bold=True, h='center')
        ws.cell(row=r,column=ci).border = _bd()
    ws.row_dimensions[r].height = 20

    return {'total_pcs':t_pcs,'total_pkgs':bundle.total_pkgs,
            'total_nw':bundle.total_nw,'total_gw':bundle.total_gw,'total_cbm':bundle.total_cbm}


# ── Sheet: 套二发票（部件明细）──────────────────────────────────────────────

def _build_set2_invoice(ws, bundle: DocumentBundle):
    ws.column_dimensions['A'].width = 18  # Mark
    ws.column_dimensions['B'].width = 14  # Unit Code
    ws.column_dimensions['C'].width = 14  # Material NO.
    ws.column_dimensions['D'].width = 14  # HS Code TH
    ws.column_dimensions['E'].width = 30  # Description
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 12  # Qty
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['I'].width = 14  # Unit Price
    ws.column_dimensions['J'].width = 16  # Amount

    r = 1
    _mc(ws,r,1,r,9, bundle.seller_name_en, bold=True, size=12)
    ws.row_dimensions[r].height = 22; r += 1
    ws.row_dimensions[r].height = 6; r += 1
    _mc(ws,r,1,r,9, bundle.seller_address_en, size=9)
    ws.row_dimensions[r].height = 16; r += 1
    for _ in range(3): ws.row_dimensions[r].height = 6; r += 1

    _mc(ws,r,1,r,1, 'Messrs：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_name_en, bold=True)
    _sc(ws,r,7, 'Invoice No.：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_no + '-TH', bold=True)  # 套二加-TH后缀区分
    ws.row_dimensions[r].height = 18; r += 1
    _mc(ws,r,1,r,1, 'Address：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_address_en)
    _sc(ws,r,7, 'Invoice Date：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_date)
    ws.row_dimensions[r].height = 26; r += 1

    for ci, h in enumerate(['Mark & No.','Unit Code','Material NO.','HS Code\n(Thailand)','Description of Goods',
        '','Quantity (PCS)','','Unit Price (USD)','Amount (USD)'], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _fn(True, 9); c.fill = _fl('D6E4F0')
        c.alignment = _al('center'); c.border = _bd()
    ws.row_dimensions[r].height = 32; r += 1

    ctnr_nos = ' / '.join(c.container_no for c in bundle.containers)
    _sc(ws,r,1, 'EF', bold=True, h='center')
    ws.row_dimensions[r].height = 16; r += 1
    _sc(ws,r,1, ctnr_nos, bold=True, h='center')
    ws.row_dimensions[r].height = 16

    last_prod = None
    for line in bundle.set2_lines:
        if line.product_code != last_prod:
            # 产品分组标题行
            fg_name = line.product_code
            _mc(ws,r,2,r,9, f'— {fg_name} —', bold=True, size=9,
                bg='F0F7FF', color='1F3864', h='center')
            ws.row_dimensions[r].height = 16; r += 1
            last_prod = line.product_code

        _sc(ws,r,2, line.product_code, h='center', size=8)
        ws.cell(row=r,column=2).border = _bd()
        _sc(ws,r,3, line.material_code, h='center', size=8)
        ws.cell(row=r,column=3).border = _bd()
        _sc(ws,r,4, line.hs_code_th, h='center', size=9)
        ws.cell(row=r,column=4).border = _bd()
        _mc(ws,r,5,r,6, line.name_en)
        _sc(ws,r,7, int(line.total_qty), h='center')
        ws.cell(row=r,column=7).border = _bd()
        _sc(ws,r,9, line.unit_price, h='center', num_fmt='#,##0.00')
        ws.cell(row=r,column=9).border = _bd()
        _sc(ws,r,10, line.total_amount, h='center', num_fmt='#,##0.00')
        ws.cell(row=r,column=10).border = _bd()
        ws.row_dimensions[r].height = 16; r += 1

    r += 1
    _mc(ws,r,2,r,9, amount_words(bundle.set2_total_amount), italic=True, size=9)
    ws.row_dimensions[r].height = 16; r += 2
    _mc(ws,r,7,r,8, 'TOTAL AMOUNT:', bold=True, h='right')
    _sc(ws,r,9, 'USD', bold=True, h='center')
    _sc(ws,r,10, bundle.set2_total_amount, bold=True, h='center', num_fmt='#,##0.00')
    ws.row_dimensions[r].height = 20


# ── Sheet: 套二PL（部件装箱单）──────────────────────────────────────────────

def _build_set2_pl(ws, bundle: DocumentBundle):
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14

    r = 1
    _mc(ws,r,1,r,10, bundle.seller_name_en, bold=True, size=12)
    ws.row_dimensions[r].height = 22; r += 1
    ws.row_dimensions[r].height = 6; r += 1
    _mc(ws,r,1,r,10, bundle.seller_address_en, size=9)
    ws.row_dimensions[r].height = 16; r += 1
    for _ in range(3): ws.row_dimensions[r].height = 6; r += 1
    _mc(ws,r,1,r,1, 'Messrs：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_name_en, bold=True)
    _sc(ws,r,7, 'Invoice No.：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_no + '-TH', bold=True)
    ws.row_dimensions[r].height = 18; r += 1
    _mc(ws,r,1,r,1, 'Address：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_address_en)
    _sc(ws,r,7, 'Invoice Date：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_date)
    ws.row_dimensions[r].height = 26; r += 1
    _mc(ws,r,3,r,10, 'PACKING  LIST (PARTS)', bold=True, size=14, h='center')
    ws.row_dimensions[r].height = 22; r += 1

    pl_hdrs = ['Mark & Container No.','Description of Goods','','Pituctures of Goods','',
               'Quantity \n(PCS)','Quantity \n(PKGS)','N.W.\n(KGS)','G.W.\n(KGS)','Measurement  (CBM)']
    for ci, h in enumerate(pl_hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _fn(True, 9); c.fill = _fl('D6E4F0')
        c.alignment = _al('center'); c.border = _bd()
    ws.row_dimensions[r].height = 36; r += 1

    ctnr_nos = ' / '.join(c.container_no for c in bundle.containers)
    _sc(ws,r,1, 'EF', bold=True, h='center')
    ws.row_dimensions[r].height = 16; r += 1
    _sc(ws,r,1, ctnr_nos, bold=True, h='center')
    ws.row_dimensions[r].height = 16

    first_cbm_row = r
    t_pcs=t_pkgs=t_nw=t_gw=t_cbm=0
    last_prod = None

    for line in bundle.set2_lines:
        if line.product_code != last_prod:
            _mc(ws,r,2,r,5, f'— {line.product_code} —', bold=True, size=9,
                bg='F0F7FF', color='1F3864', h='center')
            ws.row_dimensions[r].height = 16; r += 1
            last_prod = line.product_code

        _mc(ws,r,2,r,3, line.name_en)
        _sc(ws,r,6, int(line.total_qty), h='center'); ws.cell(row=r,column=6).border=_bd()
        pkgs = int(line.box_count)
        _sc(ws,r,7, pkgs, h='center'); ws.cell(row=r,column=7).border=_bd()
        _sc(ws,r,8, line.total_nw, h='center'); ws.cell(row=r,column=8).border=_bd()
        _sc(ws,r,9, line.total_gw, h='center'); ws.cell(row=r,column=9).border=_bd()
        ws.row_dimensions[r].height = 16
        t_pcs+=int(line.total_qty); t_pkgs+=pkgs; t_nw+=line.total_nw
        t_gw+=line.total_gw; t_cbm+=line.total_cbm; r+=1

    ws.cell(row=first_cbm_row, column=10).value = round(t_cbm, 3)
    ws.cell(row=first_cbm_row, column=10).alignment = _al('center')

    r += 1
    from num2words import num2words as n2w
    pkgs_word = n2w(t_pkgs, lang='en').upper()
    _mc(ws,r,2,r,10, f"TOTAL PACKED IN {pkgs_word} ({t_pkgs}) PACKAGES ONLY.", italic=True, size=9)
    ws.row_dimensions[r].height = 16; r += 2
    _mc(ws,r,4,r,4, 'TOTAL:', bold=True, h='right')
    for ci, val in [(6,t_pcs),(7,t_pkgs),(8,round(t_nw,2)),(9,round(t_gw,2)),(10,round(t_cbm,3))]:
        _sc(ws,r,ci, val, bold=True, h='center'); ws.cell(row=r,column=ci).border=_bd()
    ws.row_dimensions[r].height = 20


# ── Sheet: 报关单 ────────────────────────────────────────────────────────────

def _build_customs(ws, bundle: DocumentBundle):
    for col, w in [('A',14),('B',14),('C',22),('D',8),('E',8),('F',8),
                   ('G',8),('H',8),('I',12),('J',8),('K',8),('L',8),
                   ('M',10),('N',8),('O',8),('P',8),('Q',8)]:
        ws.column_dimensions[col].width = w

    def sc(r,c,v,**kw): _sc(ws,r,c,v,**kw)
    def mc(r1,c1,r2,c2,v='',**kw): _mc(ws,r1,c1,r2,c2,v,**kw)

    r = 1
    so_str = '；'.join(bundle.so_nos) if bundle.so_nos else ''
    sc(r,1, f"SO: {so_str}", size=9); ws.row_dimensions[r].height = 16; r += 1
    mc(r,1,r,17, '中华人民共和国海关出口货物报关单', bold=True, size=13, h='center')
    ws.row_dimensions[r].height = 22; r += 2

    sc(r,1,'预录入编号:',bold=True,size=8); sc(r,10,'海关编号:',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1
    for ci,lbl in [(1,'出口口岸'),(5,'备案号'),(10,'出口日期'),(14,'申报日期')]:
        sc(r,ci,lbl,bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=2

    mc(r,1,r,4, f"生产销售单位: {bundle.seller_name_cn}", bold=True, size=9)
    sc(r,5,'运输方式',bold=True,size=8); sc(r,7,'运输工具名称',bold=True,size=8)
    sc(r,11,'提运单号',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1

    mc(r,1,r,4, f"境外收货人：{bundle.buyer_name_en}", size=9)
    ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,4, f"发货单位: ({bundle.seller_tax_id})", size=9)
    sc(r,5,'贸易方式',bold=True,size=8); sc(r,9,'征免性质',bold=True,size=8); sc(r,13,'结汇方式',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,4, bundle.seller_name_cn, bold=True, size=9)
    sc(r,5,'一般贸易',size=9); sc(r,9,'一般征税',size=9); sc(r,13,'T/T',size=9)
    ws.row_dimensions[r].height=16; r+=1
    sc(r,5,'运抵国(地区)',bold=True,size=8); sc(r,9,'指运港',bold=True,size=8); sc(r,13,'境内货源地',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1
    sc(r,5,'泰国',size=9); sc(r,9,'Bangkok',size=9); sc(r,13,'广东',size=9)
    ws.row_dimensions[r].height=16; r+=1
    sc(r,4,'成交方式',bold=True,size=8); sc(r,5,'运费:',bold=True,size=8)
    sc(r,10,'保费',bold=True,size=8); sc(r,13,'杂费',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1
    sc(r,4,bundle.buyer_incoterms,size=9); ws.row_dimensions[r].height=16; r+=1
    sc(r,4,'件数',bold=True,size=8)
    sc(r,6, f"包装种类: 纸箱，木箱", size=9)
    sc(r,10,'毛重(公斤)',bold=True,size=8); sc(r,15,'净重(公斤)',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1
    sc(r,4, bundle.total_pkgs, bold=True,size=9,h='center')
    sc(r,6,'个',size=9)
    sc(r,10, bundle.total_gw, bold=True,size=9,h='center')
    sc(r,15, bundle.total_nw, bold=True,size=9,h='center')
    ws.row_dimensions[r].height=16; r+=1
    sc(r,1,'合同协议号',bold=True,size=8); sc(r,5,'随附单据',bold=True,size=8)
    ws.row_dimensions[r].height=16; r+=1
    sc(r,1,bundle.invoice_no,size=9); ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,4,'标记唛码及备注',bold=True,size=8); ws.row_dimensions[r].height=16; r+=1

    # 品牌标注
    for line in bundle.set1_lines:
        mc(r,1,r,4, line.brand_note, size=9); ws.row_dimensions[r].height=16; r+=1; break
    mc(r,1,r,4,'MADE IN CHINA        退税',size=9); ws.row_dimensions[r].height=16; r+=1

    # 商品表头
    hdrs=['项号','商品编号','商品名称、规格型号','','数量及单位','','柜号','净重(kg)',
          '最终目的国','单价','','','总价','','','币制','征免']
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=r,column=ci,value=h)
        c.font=_fn(True,8); c.alignment=_al('center'); c.border=_bd(); c.fill=_fl('D6E4F0')
    ws.row_dimensions[r].height=20; r+=1

    ctnr_str = ' / '.join(c.container_no for c in bundle.containers)
    for i, line in enumerate(bundle.set1_lines):
        _sc(ws,r,1,i+1,h='center'); ws.cell(row=r,column=1).border=_bd()
        _sc(ws,r,2,line.hs_code_cn,h='center',size=9); ws.cell(row=r,column=2).border=_bd()
        _sc(ws,r,3,line.customs_elements,size=8); ws.cell(row=r,column=3).border=_bd()
        _sc(ws,r,5,line.customs_suits,h='center'); ws.cell(row=r,column=5).border=_bd()
        _sc(ws,r,6,'套',h='center',size=9); ws.cell(row=r,column=6).border=_bd()
        _sc(ws,r,7,ctnr_str,h='center',size=8); ws.cell(row=r,column=7).border=_bd()
        _sc(ws,r,8,bundle.total_nw,h='center',size=9); ws.cell(row=r,column=8).border=_bd()
        _sc(ws,r,9,'泰国',h='center',size=9); ws.cell(row=r,column=9).border=_bd()
        _sc(ws,r,10,line.unit_price,h='center',size=9); ws.cell(row=r,column=10).border=_bd()
        _sc(ws,r,13,line.total_amount,h='center',size=9); ws.cell(row=r,column=13).border=_bd()
        _sc(ws,r,16,'USD',h='center',size=9); ws.cell(row=r,column=16).border=_bd()
        _sc(ws,r,17,'照章',h='center',size=9); ws.cell(row=r,column=17).border=_bd()
        ws.row_dimensions[r].height=22; r+=1


# ── Sheet: 合同 ─────────────────────────────────────────────────────────────

def _build_contract(ws, bundle: DocumentBundle):
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    def sc(r,c,v,**kw): _sc(ws,r,c,v,**kw)
    def mc(r1,c1,r2,c2,v='',**kw): _mc(ws,r1,c1,r2,c2,v,**kw)

    r = 1
    mc(r,1,r,6,'售货合同', bold=True, size=16, h='center'); ws.row_dimensions[r].height=28; r+=1
    mc(r,1,r,6,'Sales Contract', bold=True, size=14, h='center', color='595959'); ws.row_dimensions[r].height=22; r+=1
    sc(r,5,'合同编码',bold=True,h='right'); sc(r,6,bundle.invoice_no,bold=True)
    ws.row_dimensions[r].height=18; r+=1
    sc(r,5,'Contract No.',size=8,h='right',color='595959'); ws.row_dimensions[r].height=14; r+=1
    sc(r,5,'日期:',h='right'); sc(r,6,bundle.contract_date); ws.row_dimensions[r].height=18; r+=1
    sc(r,5,'Date:',size=8,h='right',color='595959'); ws.row_dimensions[r].height=14; r+=1

    mc(r,1,r,4, f"卖方:{bundle.seller_name_cn}",bold=True); ws.row_dimensions[r].height=18; r+=1
    mc(r,1,r,4, f"Seller: {bundle.seller_name_en}",size=9); ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,4, f"地址:{bundle.seller_address_cn}",size=9); ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,4, f"Add: {bundle.seller_address_en}",size=9); ws.row_dimensions[r].height=16; r+=2

    mc(r,1,r,4, f"买方/Buyer: {bundle.buyer_name_en}",bold=True); ws.row_dimensions[r].height=18; r+=1
    mc(r,1,r,6, f"地址/Address: {bundle.buyer_address_en}",size=9); ws.row_dimensions[r].height=28; r+=2

    mc(r,1,r,6,'双方同意按下列条款由买方购进卖方售出下列商品:',size=9); ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,6,'The Buyers agree to buy and the sellers agree to sell the following goods on the terms and conditions stated below',size=8,color='595959'); ws.row_dimensions[r].height=16; r+=1

    # 商品表头
    for ci,h in [(1,'品名及规格'),(2,'数量'),(4,'单价'),(5,'总值')]:
        _sc(ws,r,ci,h,bold=True,bg='D6E4F0',h='center'); ws.cell(row=r,column=ci).border=_bd()
    for ci in [3,6]: ws.cell(row=r,column=ci).fill=_fl('D6E4F0'); ws.cell(row=r,column=ci).border=_bd()
    ws.row_dimensions[r].height=18; r+=1

    for line in bundle.set1_lines:
        sc(r,1,f"{line.name_en} ({line.product_code})"); ws.cell(row=r,column=1).border=_bd()
        sc(r,2,line.customs_suits,h='center'); ws.cell(row=r,column=2).border=_bd()
        sc(r,3,'SET',h='center',size=9); ws.cell(row=r,column=3).border=_bd()
        sc(r,4,line.unit_price,h='center',num_fmt='#,##0.00'); ws.cell(row=r,column=4).border=_bd()
        sc(r,5,line.total_amount,h='center',num_fmt='#,##0.00'); ws.cell(row=r,column=5).border=_bd()
        sc(r,6,'USD',h='center',size=9); ws.cell(row=r,column=6).border=_bd()
        ws.row_dimensions[r].height=18; r+=1

    sc(r,1,'合计',bold=True); ws.cell(row=r,column=1).border=_bd()
    for ci in [2,3,4]: ws.cell(row=r,column=ci).border=_bd()
    sc(r,5,bundle.set1_total_amount,bold=True,h='center',num_fmt='#,##0.00'); ws.cell(row=r,column=5).border=_bd()
    sc(r,6,'',); ws.cell(row=r,column=6).border=_bd()
    ws.row_dimensions[r].height=18; r+=2

    terms = [
        f"1.合同总值/ Total Value of Contract: USD {bundle.set1_total_amount:,.2f}",
        "2.包装/Packing: 纸箱，木箱",
        "3.装运期限/Time of Shipment: ",
        f"4.装运口岸/Port of Shipment: {bundle.seller_port_loading}",
        "5.目的口岸/Port of Destination: Bangkok",
        f"6.付款条件/Payment Terms: {bundle.buyer_payment_terms}",
        f"7.成交条件/Trade Terms: {bundle.buyer_incoterms}",
        "8.装运唛头/Shipping Marks: EF",
    ]
    for t in terms:
        mc(r,1,r,6, t, size=9); ws.row_dimensions[r].height=16; r+=1

    r+=2
    sc(r,1,'卖方:',bold=True); sc(r,5,'买方:',bold=True,h='right')
    ws.row_dimensions[r].height=16; r+=1
    mc(r,1,r,2, bundle.seller_name_cn, bold=True)
    mc(r,4,r,6, bundle.buyer_name_en, bold=True, h='right')
    ws.row_dimensions[r].height=18


# ── Sheet: SI（货代补料）────────────────────────────────────────────────────

def _build_si(ws, bundle: DocumentBundle):
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10

    r = 1
    so_str = '；'.join(bundle.so_nos) if bundle.so_nos else ''
    _mc(ws,r,1,r,8, f"SO#{so_str}", bold=True, size=11)
    ws.row_dimensions[r].height = 18; r += 2

    _mc(ws,r,1,r,8, 'SHIPPING INSTRUCTION', bold=True, size=14, h='center')
    ws.row_dimensions[r].height = 24; r += 2

    sections = [
        ("1. Shipper:", f"{bundle.seller_name_en}\n{bundle.seller_address_en}"),
        ("2. Consignee:", f"{bundle.consignee_name}\n{bundle.consignee_address}"),
        ("3. Notify Party:", f"{bundle.notify_name}\n{bundle.notify_address}"),
        ("4. Port of loading:", bundle.seller_port_loading),
        ("5. Discharge port:", f"Bangkok     (以实际为准）"),
    ]
    for label, content in sections:
        _sc(ws,r,1, label, bold=True); _mc(ws,r,2,r,8, content, size=9)
        ws.row_dimensions[r].height = 18; r += 1
        if content and '\n' in content:
            ws.row_dimensions[r-1].height = 36
        ws.row_dimensions[r].height = 8; r += 1

    # 品名表头
    _sc(ws,r,1,'6. Description of Goods', bold=True)
    _sc(ws,r,2,'No.of PKGS', bold=True, h='center'); ws.cell(row=r,column=2).border=_bd()
    _sc(ws,r,3,'HS code', bold=True, h='center'); ws.cell(row=r,column=3).border=_bd()
    ws.row_dimensions[r].height = 18; r += 1

    # 套一整机品名（SI用整机归并）
    for line in bundle.set1_lines:
        _sc(ws,r,1, line.name_en)
        _sc(ws,r,2, '', h='center')  # 箱数空，汇总在货柜行
        ws.row_dimensions[r].height = 16; r += 1

    r += 1
    _sc(ws,r,1,'7. SHIPPING MARKS: EF', bold=True)
    _sc(ws,r,4,'柜重', bold=True, h='center')
    ws.row_dimensions[r].height = 18; r += 1

    # 货柜明细表头
    for ci, h in enumerate(['SO','CONTAINER NO.','SEAL NO.','PKGS','TARE','KGS','CBM','VGM','柜型'], 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _fn(True,9); c.fill=_fl('D6E4F0')
        c.alignment=_al('center'); c.border=_bd()
    ws.row_dimensions[r].height = 20; r += 1

    total_pkgs = 0
    for ctnr in bundle.containers:
        _sc(ws,r,1, ctnr.so_no, size=9)
        _sc(ws,r,2, ctnr.container_no, size=9)
        _sc(ws,r,3, ctnr.seal_no, size=9)
        rows_for_ctnr = [l for l in bundle.set2_lines if l.container_seq == ctnr.seq]
        pkgs = sum(int(l.box_count) for l in rows_for_ctnr)
        total_pkgs += pkgs
        gw  = sum(l.total_gw for l in rows_for_ctnr)
        cbm = sum(l.total_cbm for l in rows_for_ctnr)
        _sc(ws,r,4, pkgs, h='center')
        # TARE（空柜重）：40HQ约3700kg，40GP约3500kg
        tare = 3700 if '40HQ' in (ctnr.container_size or '') else 3500
        _sc(ws,r,5, tare, h='center')
        vgm = ctnr.vgm_kg or (round(gw + tare, 0))
        _sc(ws,r,6, round(gw, 0), h='center')
        _sc(ws,r,7, round(cbm, 1), h='center')
        _sc(ws,r,8, int(vgm), h='center')
        _sc(ws,r,9, ctnr.container_size or '40HQ', h='center')
        for ci in range(1, 10): ws.cell(row=r,column=ci).border=_bd()
        ws.row_dimensions[r].height = 20; r += 1

    # 合计
    _sc(ws,r,1,'TOTAL', bold=True)
    _sc(ws,r,4, total_pkgs, bold=True, h='center')
    for ci in range(1, 10): ws.cell(row=r,column=ci).border=_bd()
    ws.row_dimensions[r].height = 20


# ── 主函数：生成完整Excel文件 ────────────────────────────────────────────────

def generate_document_set(bundle: DocumentBundle, output_dir: str,
                           master_path: str = None) -> dict:
    """
    生成一套完整单据：套一 + 套二，各自一个Excel文件
    master_path: 主数据手册路径，用于读取②物料价格表N列的部件图片
    返回 {'set1': path, 'set2': path}
    """
    os.makedirs(output_dir, exist_ok=True)
    date_str = bundle.invoice_date.replace('-', '')
    base_name = f"{bundle.invoice_no}_{bundle.customer_code}_{date_str}"

    # ── 套一 Excel ────────────────────────────────────────────
    wb1 = Workbook()
    ws_inv = wb1.active; ws_inv.title = 'Inv.'
    _build_set1_invoice(ws_inv, bundle)

    ws_pl = wb1.create_sheet('PL')
    _build_set1_pl(ws_pl, bundle)

    ws_cus = wb1.create_sheet('报关单')
    _build_customs(ws_cus, bundle)

    ws_con = wb1.create_sheet('合同 ')
    _build_contract(ws_con, bundle)

    ws_si = wb1.create_sheet('SI')
    _build_si(ws_si, bundle)

    path1 = os.path.join(output_dir, f"套一_{base_name}.xlsx")
    wb1.save(path1)

    # ── 套二 Excel（含部件图片）──────────────────────────────
    # 加载图片（如果主数据路径存在）
    component_images = {}
    if master_path and os.path.exists(master_path):
        try:
            component_images = load_component_images(master_path)
        except Exception:
            pass

    wb2 = Workbook()
    ws_inv2 = wb2.active; ws_inv2.title = 'Inv.(Parts)'
    _build_set2_invoice(ws_inv2, bundle)

    ws_pl2 = wb2.create_sheet('PL(Parts)')
    _build_set2_pl_with_images(ws_pl2, bundle, component_images)

    path2 = os.path.join(output_dir, f"套二_{base_name}.xlsx")
    wb2.save(path2)

    return {'set1': path1, 'set2': path2}


# ── 图片读取工具（从主数据②物料价格表读取N列图片）────────────────────────────

def load_component_images(master_path: str) -> dict:
    """
    从主数据管理手册②物料价格表的N列读取部件图片
    返回 {material_code: image_bytes}
    """
    from openpyxl import load_workbook
    import io

    images = {}
    try:
        wb = load_workbook(master_path, data_only=True)
        # 找物料价格表Sheet
        ws = None
        for name in wb.sheetnames:
            if '物料价格' in name:
                ws = wb[name]; break
        if not ws:
            return images

        # 建立图片锚点→行号的映射
        for img in ws._images:
            try:
                # 图片锚点行（1-indexed）
                anchor = img.anchor
                if hasattr(anchor, '_from'):
                    img_row = anchor._from.row + 1  # 0-indexed→1-indexed
                elif hasattr(anchor, 'row'):
                    img_row = anchor.row
                else:
                    continue

                # 从B列（第2列）读取物料编码
                mat_code = str(ws.cell(row=img_row, column=2).value or '').strip()
                if not mat_code or mat_code in ('nan', '物料编码', '★主键2'):
                    continue

                # 读取图片数据
                raw = img._data()
                if raw and len(raw) > 100:
                    images[mat_code] = raw
            except Exception:
                continue
    except Exception:
        pass

    return images


# ── 重写套二PL（含图片）────────────────────────────────────────────────────────

def _build_set2_pl_with_images(ws, bundle, component_images: dict = None):
    """套二装箱单，支持图片嵌入"""
    import io as _io
    from openpyxl.drawing.image import Image as XLImage

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 14   # 图片列
    ws.column_dimensions['E'].width = 4
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14

    r = 1
    _mc(ws,r,1,r,10, bundle.seller_name_en, bold=True, size=12)
    ws.row_dimensions[r].height=22; r+=1
    ws.row_dimensions[r].height=6; r+=1
    _mc(ws,r,1,r,10, bundle.seller_address_en, size=9)
    ws.row_dimensions[r].height=16; r+=1
    for _ in range(3): ws.row_dimensions[r].height=6; r+=1
    _mc(ws,r,1,r,1, 'Messrs：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_name_en, bold=True)
    _sc(ws,r,7, 'Invoice No.：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_no + '-TH', bold=True)
    ws.row_dimensions[r].height=18; r+=1
    _mc(ws,r,1,r,1, 'Address：', bold=True)
    _mc(ws,r,2,r,5, bundle.buyer_address_en)
    _sc(ws,r,7, 'Invoice Date：', bold=True, h='right')
    _sc(ws,r,8, bundle.invoice_date)
    ws.row_dimensions[r].height=26; r+=1
    _mc(ws,r,3,r,10, 'PACKING  LIST (PARTS)', bold=True, size=14, h='center')
    ws.row_dimensions[r].height=22; r+=1

    pl_hdrs = ['Mark & Container No.','Description of Goods','',
               'Pituctures\nof Goods','',
               'Quantity \n(PCS)','Quantity \n(PKGS)','N.W.\n(KGS)','G.W.\n(KGS)','Measurement  (CBM)']
    for ci, h in enumerate(pl_hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _fn(True, 9); c.fill = _fl('D6E4F0')
        c.alignment = _al('center'); c.border = _bd()
    ws.row_dimensions[r].height=36; r+=1

    ctnr_nos = ' / '.join(c.container_no for c in bundle.containers)
    _sc(ws,r,1, 'EF', bold=True, h='center')
    ws.row_dimensions[r].height=16; r+=1
    _sc(ws,r,1, ctnr_nos, bold=True, h='center')
    ws.row_dimensions[r].height=16

    first_cbm_row = r
    t_pcs=t_pkgs=t_nw=t_gw=t_cbm=0
    last_prod = None
    IMG_ROW_H = 60  # 有图片的行高

    for line in bundle.set2_lines:
        if line.product_code != last_prod:
            _mc(ws,r,2,r,5, f'— {line.product_code} —', bold=True, size=9,
                bg='F0F7FF', color='1F3864', h='center')
            ws.row_dimensions[r].height=16; r+=1
            last_prod = line.product_code

        _mc(ws,r,2,r,3, line.name_en)
        _sc(ws,r,6, int(line.total_qty), h='center'); ws.cell(row=r,column=6).border=_bd()
        pkgs=int(line.box_count)
        _sc(ws,r,7, pkgs, h='center'); ws.cell(row=r,column=7).border=_bd()
        _sc(ws,r,8, line.total_nw, h='center'); ws.cell(row=r,column=8).border=_bd()
        _sc(ws,r,9, line.total_gw, h='center'); ws.cell(row=r,column=9).border=_bd()

        # 嵌入图片到D列
        has_img = False
        if component_images and line.material_code in component_images:
            try:
                img_bytes = component_images[line.material_code]
                img_stream = _io.BytesIO(img_bytes)
                xl_img = XLImage(img_stream)
                # 适配单元格大小
                xl_img.width = 50
                xl_img.height = 50
                col_letter = 'D'
                ws.add_image(xl_img, f'{col_letter}{r}')
                has_img = True
            except Exception:
                pass

        ws.row_dimensions[r].height = IMG_ROW_H if has_img else 16
        t_pcs+=int(line.total_qty); t_pkgs+=pkgs
        t_nw+=line.total_nw; t_gw+=line.total_gw; t_cbm+=line.total_cbm
        r+=1

    ws.cell(row=first_cbm_row, column=10).value = round(t_cbm, 3)
    ws.cell(row=first_cbm_row, column=10).alignment = _al('center')

    r+=1
    try:
        from num2words import num2words as n2w
        pkgs_word = n2w(t_pkgs, lang='en').upper()
    except:
        pkgs_word = str(t_pkgs)
    _mc(ws,r,2,r,10, f"TOTAL PACKED IN {pkgs_word} ({t_pkgs}) PACKAGES ONLY.",
        italic=True, size=9)
    ws.row_dimensions[r].height=16; r+=2
    _mc(ws,r,4,r,4, 'TOTAL:', bold=True, h='right')
    for ci, val in [(6,t_pcs),(7,t_pkgs),(8,round(t_nw,2)),(9,round(t_gw,2)),(10,round(t_cbm,3))]:
        _sc(ws,r,ci, val, bold=True, h='center'); ws.cell(row=r,column=ci).border=_bd()
    ws.row_dimensions[r].height=20
