"""
output_v6.py — 统一设计语言，完整边框，按官方报关单规范填写
配色：黑色文字 + 深蓝(1F3864)表头 + 浅蓝(D6E4F0)背景，仅3色
字体：统一 Arial
"""
import os, io, math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ── 设计系统常量 ──────────────────────────────────────────────────────────
NAVY = '1F3864'      # 深蓝：主标题、表头底色、强调
LBLUE = 'D6E4F0'      # 浅蓝：分组小标题底色
ZEBRA = 'F4F6F9'      # 极淡蓝灰：斑马纹底色（数据行间隔）
GRAY = '595959'       # 灰：辅助说明文字
BLACK = '000000'      # 黑：正文
FONT = 'Arial'

def F(bold=False, size=10, color=BLACK, italic=False):
    return Font(name=FONT, bold=bold, size=size, color=color, italic=italic)

def FILL(color):
    return PatternFill('solid', start_color=color, fgColor=color)

def AL(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

THIN = Side(border_style='thin', color='000000')
NONE_SIDE = Side()

def BD(t=False, b=False, l=False, r=False):
    return Border(
        top=THIN if t else NONE_SIDE, bottom=THIN if b else NONE_SIDE,
        left=THIN if l else NONE_SIDE, right=THIN if r else NONE_SIDE
    )

def FULL(): return BD(True, True, True, True)

def put(ws, r, c, v='', bold=False, size=9, color=BLACK, h='left', v_='center',
        wrap=True, num_fmt=None, bg=None, border=None, italic=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = F(bold, size, color, italic)
    cell.alignment = AL(h, v_, wrap)
    if bg: cell.fill = FILL(bg)
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell

def mc(ws, r1, c1, r2, c2):
    try: ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except: pass

def mc_put(ws, r1, c1, r2, c2, v='', **kw):
    mc(ws, r1, c1, r2, c2)
    return put(ws, r1, c1, v, **kw)

def fill_border(ws, r1, c1, r2, c2, t=True, b=True, l=True, r=True):
    """给一个矩形区域整体描边（外框），内部空白不留缝"""
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1):
            ws.cell(rr, cc).border = BD(
                t=(rr==r1 and t), b=(rr==r2 and b),
                l=(cc==c1 and l), r=(cc==c2 and r)
            )

def block_full(ws, r1, c1, r2, c2):
    """给区域每个格子都加全边框（用于数据表格区，不只是外框）"""
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1):
            ws.cell(rr, cc).border = FULL()


# ── 工具函数 ──────────────────────────────────────────────────────────────
try:
    from num2words import num2words as _n2w
    def amount_en(amt):
        d = int(amt); c = round((amt - d) * 100)
        w = _n2w(d, lang='en').upper()
        cents = f' AND CENTS {_n2w(c, lang="en").upper()}' if c else ''
        return f"SAY TOTAL U.S. DOLLARS {w}{cents} ONLY."
    def pkgs_en(n): return _n2w(int(n), lang='en').upper()
except:
    def amount_en(amt): return f"SAY TOTAL U.S. DOLLARS {amt:,.2f} ONLY."
    def pkgs_en(n): return str(n)

def compress_image(raw_bytes, size=100, quality=70):
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(raw_bytes)).convert('RGB')
        img.thumbnail((size, size), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format='JPEG', quality=quality, optimize=True)
        return out.getvalue()
    except:
        return raw_bytes

def load_component_images(master_path):
    images = {}
    if not master_path or not os.path.exists(master_path): return images
    try:
        wb = load_workbook(master_path, data_only=True)
        for name in wb.sheetnames:
            if '物料价格' in name:
                ws = wb[name]
                for img in ws._images:
                    try:
                        a = img.anchor
                        if not hasattr(a, '_from'): continue
                        row = a._from.row + 1
                        mat = str(ws.cell(row=row, column=2).value or '').strip()
                        if not mat or mat in ('nan','物料编码'): continue
                        raw = img._data()
                        if raw and len(raw) > 100: images[mat] = raw
                    except: continue
                break
    except: pass
    return images


# ════════════════════════════════════════════════════════════
# 通用抬头（Shipper / Buyer / Invoice No&Date）
# 统一样式：所有Sheet抬头区一致，标题用NAVY加粗，分隔线用thin
# ════════════════════════════════════════════════════════════
def write_header(ws, bundle, total_cols, title, party_label='Buyer',
                  party_name=None, party_addr=None):
    """
    排版层级（从粗到细）：
    公司名 16pt bold  > 标题 13pt bold+letterspace效果(用空格模拟) >
    标签 9pt bold navy > 数据 10pt regular > 说明/地址 8.5pt gray
    """
    name = party_name if party_name else bundle.buyer_name_en
    addr = party_addr if party_addr else bundle.buyer_address_en

    # R1-2: Shipper（卖方）— 最高视觉权重
    mc_put(ws,1,1,2,total_cols, bundle.seller_name_en, bold=True, size=16, h='center')
    ws.row_dimensions[1].height=22; ws.row_dimensions[2].height=10
    # R3: 地址 — 小字灰色，明显次于公司名
    mc_put(ws,3,1,3,total_cols, bundle.seller_address_en, size=8.5, color=GRAY, h='center')
    ws.row_dimensions[3].height=13
    # R4: 留白呼吸（不加分隔线，靠空间分隔层级）
    ws.row_dimensions[4].height=8

    # R5: Buyer/Consignee 标签 + Invoice No.
    mc(ws,5,1,5,4)
    put(ws,5,1,f'{party_label.upper()}', bold=True, size=8.5, color=NAVY)
    put(ws,5,5,'INVOICE NO.', bold=True, size=8.5, color=NAVY, h='right')
    mc(ws,5,6,5,total_cols)
    put(ws,5,6, bundle.invoice_no, bold=True, size=10)
    ws.row_dimensions[5].height=13

    # R6: 名称（较大）+ Invoice Date
    mc(ws,6,1,6,4)
    put(ws,6,1, name, bold=True, size=11)
    put(ws,6,5,'INVOICE DATE', bold=True, size=8.5, color=NAVY, h='right')
    mc(ws,6,6,6,total_cols)
    put(ws,6,6, bundle.invoice_date, size=10)
    ws.row_dimensions[6].height=17

    # R7: 地址
    mc(ws,7,1,7,4)
    put(ws,7,1, addr, size=8.5, color=GRAY, wrap=True)
    ws.row_dimensions[7].height=20

    # R8: 呼吸留白
    ws.row_dimensions[8].height=10

    # R9: 文件标题 — 全文档最强视觉锚点，居中+宽字距感(用Unicode间距模拟)
    spaced_title = '  '.join(title)  # 字母间加宽间距，类似letter-spacing
    mc_put(ws,9,1,9,total_cols, spaced_title, bold=True, size=15, color=NAVY, h='center')
    ws.row_dimensions[9].height=26

    # R10: 标题下细线（仅此一条强调线，全篇唯一的硬分隔）
    fill_border(ws,10,1,10,total_cols, t=False,b=True,l=False,r=False)
    for c in range(1,total_cols+1):
        ws.cell(10,c).border = Border(bottom=Side(border_style='medium', color=NAVY))
    ws.row_dimensions[10].height=4

    # R11: 货柜号 — 小字说明，留白呼吸后再进入表格
    ctnr = ' / '.join(c.container_no for c in bundle.containers if c.container_no)
    mc_put(ws,11,1,11,total_cols, f'CONTAINER NO.   {ctnr}', size=8.5, color=GRAY, h='center')
    ws.row_dimensions[11].height=16
    ws.row_dimensions[12].height=6  # 表头前留白

    return 13  # 下一行起始行号（表头行，留了一行呼吸空间）


# ════════════════════════════════════════════════════════════
# 套二 Commercial Invoice
# ════════════════════════════════════════════════════════════
def build_inv2(ws, bundle, imgs=None):
    """
    排版改造：
    - 去掉每行重复的 Mark(EF) 和 Unit Code(Product Code) —— 这两项已在分组标题行出现一次，
      行内重复15次纯属视觉噪音，删掉后数据行只保留真正变化的信息
    - 列收缩为5列：Material NO. / Description / Qty / Unit Price / Amount
    - 斑马纹：每隔一行用极淡蓝灰底色，替代密集黑白对比，降低视觉疲劳
    - 数字列右对齐+千分位，建立清晰的纵向对齐轴
    """
    NC = 5
    for col,w in [('A',18),('B',38),('C',12),('D',15),('E',17)]:
        ws.column_dimensions[col].width = w

    next_r = write_header(ws, bundle, NC, 'COMMERCIAL INVOICE')

    r = next_r
    hdrs = [(1,'MATERIAL NO.'),(2,'DESCRIPTION OF GOODS'),(3,'QTY\n(PCS)'),
            (4,'UNIT PRICE\n(USD)'),(5,'AMOUNT\n(USD)')]
    for c,h in hdrs:
        cell = ws.cell(r,c,h)
        cell.font = F(True,9,'FFFFFF'); cell.fill = FILL(NAVY)
        cell.border = FULL(); cell.alignment = AL('center')
    ws.row_dimensions[r].height = 26
    r += 1

    last_prod = None
    zebra_idx = 0
    for line in bundle.set2_lines:
        if line.product_code != last_prod:
            # 分组标题行：EF唛头 + Product Code 只在这里出现一次
            mc_put(ws,r,1,r,NC, f'EF   ·   PRODUCT {line.product_code}',
                   bold=True, size=9, color=NAVY, bg=LBLUE, border=FULL(), h='left')
            ws.row_dimensions[r].height=18; r+=1
            last_prod = line.product_code
            zebra_idx = 0  # 每组重新起算斑马纹

        bg = ZEBRA if zebra_idx % 2 == 1 else None
        for c in range(1, NC+1):
            ws.cell(r,c).border = FULL()
            if bg: ws.cell(r,c).fill = FILL(bg)
        put(ws,r,1,line.material_code,size=9,h='center',bg=bg)
        put(ws,r,2,line.name_en,size=9.5,bg=bg)
        put(ws,r,3,int(line.total_qty),h='right',bg=bg,num_fmt='#,##0')
        put(ws,r,4,line.unit_price,h='right',num_fmt='#,##0.00',bg=bg)
        put(ws,r,5,line.total_amount,h='right',num_fmt='#,##0.00',bg=bg)
        ws.row_dimensions[r].height=17
        zebra_idx += 1
        r += 1

    ws.row_dimensions[r].height=8; r+=1  # 呼吸留白后接金额大写
    mc_put(ws,r,1,r,NC, amount_en(bundle.set2_total_amount), italic=True, size=8.5, color=GRAY)
    ws.row_dimensions[r].height=14; r+=1
    ws.row_dimensions[r].height=6; r+=1

    mc(ws,r,1,r,3)
    put(ws,r,1,'TOTAL AMOUNT',bold=True,size=11,color=NAVY,h='right',border=FULL())
    put(ws,r,4,'USD',bold=True,size=11,h='center',border=FULL(),bg=LBLUE)
    put(ws,r,5,bundle.set2_total_amount,bold=True,size=11,h='right',
        num_fmt='#,##0.00',border=FULL(),bg=LBLUE)
    ws.row_dimensions[r].height=24


# ════════════════════════════════════════════════════════════
# 套二 Packing List
# ════════════════════════════════════════════════════════════
def build_pl2(ws, bundle, imgs=None):
    """
    同样去掉行内重复的 EF 唛头（移到分组标题行），数据行更干净
    斑马纹 + 数字右对齐，与 CI 保持一致的排版语言
    """
    NC = 7
    IMG_PX = 98
    IMG_ROW_H = 80
    for col,w in [('A',32),('B',14),('C',12),('D',12),('E',11),('F',11),('G',13)]:
        ws.column_dimensions[col].width = w

    next_r = write_header(ws, bundle, NC, 'PACKING LIST')

    r = next_r
    hdrs = [(1,'DESCRIPTION OF GOODS'),(2,'PICTURE'),(3,'QTY\n(PCS)'),
            (4,'QTY\n(PKGS)'),(5,'N.W.\n(KGS)'),(6,'G.W.\n(KGS)'),(7,'CBM')]
    for c,h in hdrs:
        cell = ws.cell(r,c,h)
        cell.font = F(True,9,'FFFFFF'); cell.fill = FILL(NAVY)
        cell.border = FULL(); cell.alignment = AL('center')
    ws.row_dimensions[r].height = 26
    r += 1

    t_pcs=t_pkgs=t_nw=t_gw=t_cbm=0
    last_prod=None
    cbm_first_r=None
    zebra_idx = 0

    for line in bundle.set2_lines:
        if line.product_code != last_prod:
            mc_put(ws,r,1,r,NC, f'EF   ·   PRODUCT {line.product_code}',
                   bold=True, size=9, color=NAVY, bg=LBLUE, border=FULL(), h='left')
            ws.row_dimensions[r].height=18; r+=1
            last_prod = line.product_code
            zebra_idx = 0

        bg = ZEBRA if zebra_idx % 2 == 1 else None
        for c in range(1, NC+1):
            ws.cell(r,c).border = FULL()
            if bg: ws.cell(r,c).fill = FILL(bg)
        put(ws,r,1,line.name_en,size=9.5,bg=bg)

        has_img=False
        if imgs and line.material_code in imgs:
            try:
                raw = compress_image(imgs[line.material_code], size=100, quality=70)
                xl_img = XLImage(io.BytesIO(raw))
                xl_img.width = IMG_PX; xl_img.height = 78
                ws.add_image(xl_img, f'B{r}')
                has_img = True
            except: pass

        if line.box_count and line.box_count > 0:
            calc_pkgs = int(line.box_count)
        elif getattr(line,'qty_per_box',0) and line.qty_per_box > 0:
            calc_pkgs = math.ceil(line.total_qty / line.qty_per_box)
        else:
            calc_pkgs = 0

        put(ws,r,3,int(line.total_qty),h='right',bg=bg,num_fmt='#,##0')
        put(ws,r,4,calc_pkgs,h='right',bg=bg,num_fmt='#,##0')
        nw_v = round(line.total_nw,2) if line.total_nw else ''
        gw_v = round(line.total_gw,2) if line.total_gw else ''
        put(ws,r,5,nw_v,h='right',num_fmt='#,##0.00',bg=bg)
        put(ws,r,6,gw_v,h='right',num_fmt='#,##0.00',bg=bg)
        if cbm_first_r is None: cbm_first_r = r

        ws.row_dimensions[r].height = IMG_ROW_H if has_img else 18
        t_pcs+=int(line.total_qty); t_pkgs+=calc_pkgs
        t_nw+=line.total_nw; t_gw+=line.total_gw; t_cbm+=line.total_cbm
        zebra_idx += 1
        r+=1

    if cbm_first_r:
        ws.cell(cbm_first_r,7,round(t_cbm,3))
        ws.cell(cbm_first_r,7).number_format='#,##0.000'
        ws.cell(cbm_first_r,7).alignment = AL('right')

    ws.row_dimensions[r].height=8; r+=1

    mc(ws,r,1,r,2)
    put(ws,r,1,'TOTAL',bold=True,size=10,color=NAVY,h='right',border=FULL(),bg=LBLUE)
    for c in [3,4,5,6,7]: ws.cell(r,c).border=FULL(); ws.cell(r,c).fill=FILL(LBLUE)
    put(ws,r,3,t_pcs,bold=True,h='right',num_fmt='#,##0',bg=LBLUE)
    put(ws,r,4,t_pkgs,bold=True,h='right',num_fmt='#,##0',bg=LBLUE)
    put(ws,r,5,round(t_nw,2) if t_nw else '',bold=True,h='right',num_fmt='#,##0.00',bg=LBLUE)
    put(ws,r,6,round(t_gw,2) if t_gw else '',bold=True,h='right',num_fmt='#,##0.00',bg=LBLUE)
    put(ws,r,7,round(t_cbm,3),bold=True,h='right',num_fmt='#,##0.000',bg=LBLUE)
    ws.row_dimensions[r].height=22; r+=1

    for c,u in [(3,'PCS'),(4,'PKGS'),(5,'KGS'),(6,'KGS'),(7,'CBM')]:
        put(ws,r,c,u,size=7.5,color=GRAY,h='right',italic=True)
    ws.row_dimensions[r].height=12


# ════════════════════════════════════════════════════════════
# 套一 Commercial Invoice
# ════════════════════════════════════════════════════════════
def build_inv1(ws, bundle):
    NC = 4
    for col,w in [('A',46),('B',12),('C',16),('D',18)]:
        ws.column_dimensions[col].width = w

    next_r = write_header(ws, bundle, NC, 'COMMERCIAL INVOICE')

    r = next_r
    hdrs = [(1,'DESCRIPTION OF GOODS'),(2,'QTY\n(SETS)'),
            (3,'UNIT PRICE\n(USD)'),(4,'AMOUNT\n(USD)')]
    for c,h in hdrs:
        cell = ws.cell(r,c,h)
        cell.font = F(True,9,'FFFFFF'); cell.fill = FILL(NAVY)
        cell.border = FULL(); cell.alignment = AL('center')
    ws.row_dimensions[r].height = 26
    r += 1

    t_sets = 0
    zebra_idx = 0
    for line in bundle.set1_lines:
        bg = ZEBRA if zebra_idx % 2 == 1 else None
        for c in range(1, NC+1):
            ws.cell(r,c).border = FULL()
            if bg: ws.cell(r,c).fill = FILL(bg)
        put(ws,r,1,line.name_en,size=10,bg=bg)
        put(ws,r,2,int(line.customs_suits),h='right',num_fmt='#,##0',bg=bg)
        put(ws,r,3,line.unit_price,h='right',num_fmt='#,##0.00',bg=bg)
        put(ws,r,4,line.total_amount,h='right',num_fmt='#,##0.00',bg=bg)
        t_sets += line.customs_suits
        ws.row_dimensions[r].height=20; zebra_idx+=1; r+=1

    ws.row_dimensions[r].height=8; r+=1
    mc_put(ws,r,1,r,NC, amount_en(bundle.set1_total_amount), italic=True, size=8.5, color=GRAY)
    ws.row_dimensions[r].height=14; r+=1
    ws.row_dimensions[r].height=6; r+=1

    mc(ws,r,1,r,2)
    put(ws,r,1,'TOTAL AMOUNT',bold=True,size=11,color=NAVY,h='right',border=FULL(),bg=LBLUE)
    put(ws,r,3,'USD',bold=True,size=11,h='center',border=FULL(),bg=LBLUE)
    put(ws,r,4,bundle.set1_total_amount,bold=True,size=11,h='right',
        num_fmt='#,##0.00',border=FULL(),bg=LBLUE)
    ws.row_dimensions[r].height=24


# ════════════════════════════════════════════════════════════
# 套一 Packing List
# Mark列：第一行EF，后续行=货柜号
# PKGS = SETS（整机每套1箱）
# ════════════════════════════════════════════════════════════
def build_pl1(ws, bundle):
    NC = 6
    for col,w in [('A',40),('B',12),('C',12),('D',11),('E',11),('F',12)]:
        ws.column_dimensions[col].width = w

    next_r = write_header(ws, bundle, NC, 'PACKING LIST')

    r = next_r
    hdrs = [(1,'DESCRIPTION OF GOODS'),(2,'QTY\n(SETS)'),(3,'QTY\n(PKGS)'),
            (4,'N.W.\n(KGS)'),(5,'G.W.\n(KGS)'),(6,'CBM')]
    for c,h in hdrs:
        cell = ws.cell(r,c,h)
        cell.font = F(True,9,'FFFFFF'); cell.fill = FILL(NAVY)
        cell.border = FULL(); cell.alignment = AL('center')
    ws.row_dimensions[r].height = 26
    r += 1

    ctnr_list = [c.container_no for c in bundle.containers if c.container_no]
    t_sets = 0
    zebra_idx = 0
    for li, line in enumerate(bundle.set1_lines):
        sets_val = int(line.customs_suits)
        bg = ZEBRA if zebra_idx % 2 == 1 else None
        for c in range(1, NC+1):
            ws.cell(r,c).border = FULL()
            if bg: ws.cell(r,c).fill = FILL(bg)
        put(ws,r,1,line.name_en,size=10,bg=bg)
        put(ws,r,2,sets_val,h='right',num_fmt='#,##0',bg=bg)
        put(ws,r,3,sets_val,h='right',num_fmt='#,##0',bg=bg)
        if li == 0:
            nw_v = round(bundle.total_nw,2) if bundle.total_nw else ''
            gw_v = round(bundle.total_gw,2) if bundle.total_gw else ''
            put(ws,r,4,nw_v,h='right',num_fmt='#,##0.00',bg=bg)
            put(ws,r,5,gw_v,h='right',num_fmt='#,##0.00',bg=bg)
            put(ws,r,6,round(bundle.total_cbm,3),h='right',num_fmt='#,##0.000',bg=bg)
        t_sets += sets_val
        ws.row_dimensions[r].height=20; zebra_idx+=1; r+=1

    # 货柜号行 — 小字弱化处理，明显是辅助信息而非数据行
    if ctnr_list:
        mc(ws,r,1,r,NC)
        put(ws,r,1,f"EF   ·   Container No.:  {'  /  '.join(ctnr_list)}",
            size=8.5, color=GRAY, italic=True, border=FULL(), bg=LBLUE)
        ws.row_dimensions[r].height=16; r+=1

    ws.row_dimensions[r].height=8; r+=1

    mc(ws,r,1,r,1)
    put(ws,r,1,'TOTAL',bold=True,size=10,color=NAVY,h='right',border=FULL(),bg=LBLUE)
    for c in [2,3,4,5,6]: ws.cell(r,c).border=FULL(); ws.cell(r,c).fill=FILL(LBLUE)
    put(ws,r,2,t_sets,bold=True,h='right',num_fmt='#,##0',bg=LBLUE)
    put(ws,r,3,t_sets,bold=True,h='right',num_fmt='#,##0',bg=LBLUE)
    put(ws,r,4,round(bundle.total_nw,2) if bundle.total_nw else '',bold=True,h='right',num_fmt='#,##0.00',bg=LBLUE)
    put(ws,r,5,round(bundle.total_gw,2) if bundle.total_gw else '',bold=True,h='right',num_fmt='#,##0.00',bg=LBLUE)
    put(ws,r,6,round(bundle.total_cbm,3),bold=True,h='right',num_fmt='#,##0.000',bg=LBLUE)
    ws.row_dimensions[r].height=22; r+=1

    for c,u in [(2,'SETS'),(3,'PKGS'),(4,'KGS'),(5,'KGS'),(6,'CBM')]:
        put(ws,r,c,u,size=7.5,color=GRAY,h='right',italic=True)
    ws.row_dimensions[r].height=12


# ════════════════════════════════════════════════════════════
# 报关单 — 按海关官方规范填写
# 能确定的栏目：填入实际值
# 海关/货代专属栏目（出口口岸/出口日期/申报日期/运输工具/提运单号/
#                     运费保费杂费/许可证号）：留空，由货代补充
# 统一边框：所有栏目格满边框，不留缺口
# ════════════════════════════════════════════════════════════
def build_customs(ws, bundle):
    """
    报关单 — 按海关总署2018年第61号/2019年第18号公告现行规范填写
    主要修正：
      境内发货人（原：收发货人/经营单位）
      出境关别（原：出口口岸）
      运输工具名称及航次号（原：运输工具名称）
      随附单证及编号（原：随附单证）
      新增：境外收货人、离境口岸、原产国(地区)
      删除：录入员/录入单位（2018年起已废止）
    """
    for col,w in [('A',4.5),('B',18),('C',16),('D',9),('E',9),
                   ('F',5),('G',14),('H',13),('I',9),('J',3),
                   ('K',10),('L',5),('M',4),('N',7),('O',6),('P',7)]:
        ws.column_dimensions[col].width = w

    so_str   = '；'.join(dict.fromkeys(bundle.so_nos)) if bundle.so_nos else ''
    ctnr_str = '  '.join(c.container_no for c in bundle.containers if c.container_no)
    tax      = bundle.seller_tax_id or ''
    name_cn  = bundle.seller_name_cn or bundle.seller_name_en
    gw_val   = round(bundle.total_gw,0) if bundle.total_gw else 0
    nw_val   = round(bundle.total_nw,0) if bundle.total_nw else 0

    def label_value_block(r1,c1,r2,c2,label,value=None,pending=False):
        mc(ws,r1,c1,r2,c2)
        if pending:
            content = f"{label}\n（待补充）"; color = GRAY
        elif value is not None and value != '':
            content = f"{label}\n{value}"; color = BLACK
        else:
            content = label; color = BLACK
        cell = ws.cell(r1,c1,content)
        cell.font = F(False,8,color,italic=pending)
        cell.alignment = AL('center','top',True)
        block_full(ws,r1,c1,r2,c2)

    # R1: SO号
    put(ws,1,1,'SO:',size=8,bold=True); put(ws,1,2,so_str,size=8)
    ws.row_dimensions[1].height = 14

    # R2: 标题
    mc_put(ws,2,1,2,16,'中华人民共和国海关出口货物报关单',bold=True,size=13,h='center')
    ws.row_dimensions[2].height = 21; ws.row_dimensions[3].height = 6

    # R4: 预录入/海关编号（系统生成，说明性文字）
    mc(ws,4,3,4,8); put(ws,4,3,'预录入编号: （海关/报关行系统生成）',size=7,color=GRAY,italic=True)
    mc(ws,4,10,4,16); put(ws,4,10,'海关编号: （海关接受申报后给予）',size=7,color=GRAY,italic=True)
    ws.row_dimensions[4].height = 13

    # R5-6: 出境关别/备案号/出口日期/申报日期
    label_value_block(5,1,6,4,'出境关别',pending=True)
    label_value_block(5,5,6,10,'备案号','一般贸易免予填报')
    label_value_block(5,11,6,13,'出口日期',pending=True)
    label_value_block(5,14,6,16,'申报日期',pending=True)
    ws.row_dimensions[5].height=12.8; ws.row_dimensions[6].height=14

    # R7-8: 境内发货人/运输方式/运输工具名称及航次号/提运单号
    label_value_block(7,1,8,4,'境内发货人', f'{name_cn}\n统一社会信用代码:{tax}')
    label_value_block(7,5,8,6,'运输方式','水路运输')
    label_value_block(7,7,8,11,'运输工具名称及航次号',pending=True)
    label_value_block(7,12,8,16,'提运单号(B/L NO.)',pending=True)
    ws.row_dimensions[7].height=12; ws.row_dimensions[8].height=28

    # R9-10: 境外收货人/生产销售单位
    label_value_block(9,1,10,4,'境外收货人', bundle.consignee_name)
    label_value_block(9,5,10,9,'生产销售单位', name_cn)
    ws.row_dimensions[9].height=12; ws.row_dimensions[10].height=16

    # R11-12: 监管方式/征免性质/结汇方式/许可证号
    label_value_block(11,1,12,4,'监管方式（贸易方式）','一般贸易 0110')
    label_value_block(11,5,12,8,'征免性质','一般征税')
    label_value_block(11,9,12,12,'结汇方式','T/T')
    label_value_block(11,13,12,16,'许可证号','无')
    ws.row_dimensions[11].height=12; ws.row_dimensions[12].height=14

    # R13-14: 运抵国/离境口岸/指运港/境内货源地
    label_value_block(13,1,14,4,'运抵国(地区)','泰国 THAILAND')
    label_value_block(13,5,14,8,'离境口岸',pending=True)
    label_value_block(13,9,14,12,'指运港', bundle.port_discharge or '曼谷 BANGKOK')
    label_value_block(13,13,14,16,'境内货源地','广东省中山市')
    ws.row_dimensions[13].height=12; ws.row_dimensions[14].height=14

    # R15: 成交方式/运费/保费/杂费
    label_value_block(15,1,15,3,'本栏目空白','')
    label_value_block(15,4,15,5,'成交方式', bundle.buyer_incoterms or 'FOB')
    label_value_block(15,6,15,9,'运费',pending=True)
    label_value_block(15,10,15,12,'保费',pending=True)
    label_value_block(15,13,15,16,'杂费',pending=True)
    ws.row_dimensions[15].height=24

    # R16-17: 合同协议号/件数/包装种类/毛重/净重
    label_value_block(16,1,17,3,'合同协议号', bundle.invoice_no)
    label_value_block(16,4,17,5,'件数', f'{bundle.total_pkgs} 件')
    label_value_block(16,6,17,9,'包装种类','纸箱（1套=1件）')
    label_value_block(16,10,17,13,'毛重(公斤)', f'{gw_val:,.0f}' if gw_val else '（待补充）')
    label_value_block(16,14,17,16,'净重(公斤)', f'{nw_val:,.0f}' if nw_val else '（待补充）')
    ws.row_dimensions[16].height=12; ws.row_dimensions[17].height=24

    # R18-19: 集装箱号/随附单证及编号/原产国(地区)
    ctnr_display = ctnr_str.replace('  ', '\n')
    mc(ws,18,1,19,5)
    c18 = ws.cell(18,1, f'集装箱号\n{ctnr_display}')
    c18.font = F(False,8); c18.alignment = AL('center','top',True); block_full(ws,18,1,19,5)

    mc(ws,18,6,19,11)
    c18b = ws.cell(18,6,'随附单证及编号\n合同、发票、装箱单')
    c18b.font = F(False,8); c18b.alignment = AL('center','top',True); block_full(ws,18,6,19,11)

    mc(ws,18,12,19,16)
    c18c = ws.cell(18,12,'原产国(地区)\n中国 CHINA')
    c18c.font = F(False,8); c18c.alignment = AL('center','top',True); block_full(ws,18,12,19,16)
    ws.row_dimensions[18].height=14; ws.row_dimensions[19].height=22

    # R20: 生产厂家
    mc_put(ws,20,1,20,16, f'生产厂家：{name_cn}', size=8, border=FULL())
    ws.row_dimensions[20].height=15

    # R21-23: 标记唛码/品牌 / MADE IN CHINA / 其他事项确认
    brand = bundle.set1_lines[0].brand_note if bundle.set1_lines else '无品牌'
    mc_put(ws,21,1,21,16, f'标记唛码及备注：EF  |  包装种类：纸箱  |  品牌类型：{brand}',
           size=8, border=FULL())
    ws.row_dimensions[21].height=15
    mc_put(ws,22,1,22,16,'MADE IN CHINA   出口享惠情况：不享惠/一般原产地证',
           size=8, border=FULL())
    ws.row_dimensions[22].height=15
    mc_put(ws,23,1,23,16,'特殊关系确认：否    价格影响确认：（不适用）    支付特许权使用费确认：否',
           size=8, border=FULL())
    ws.row_dimensions[23].height=15

    # R24: 商品明细表头
    HDR_R = 24
    mc(ws,HDR_R,3,HDR_R,4); mc(ws,HDR_R,5,HDR_R,6)
    mc(ws,HDR_R,9,HDR_R,10); mc(ws,HDR_R,12,HDR_R,14)
    hdr_map = {1:'项号',2:'商品编号',3:'商品名称、规格型号',
               5:'数量及单位',7:'柜号',8:'净重(kg)',
               9:'最终目的国(地区)',11:'单价',12:'总价',15:'币制',16:'征免'}
    for c in range(1,17):
        cell = ws.cell(HDR_R,c)
        cell.font = F(True,8,'FFFFFF'); cell.fill = FILL(NAVY)
        cell.alignment = AL('center'); cell.border = FULL()
        if c in hdr_map: cell.value = hdr_map[c]
    ws.row_dimensions[HDR_R].height = 20

    # R25+: 商品行
    base_r = HDR_R + 1
    for i, line in enumerate(bundle.set1_lines):
        r = base_r + i
        ws.row_dimensions[r].height = 75
        mc(ws,r,3,r,4); mc(ws,r,5,r,6); mc(ws,r,9,r,10); mc(ws,r,12,r,14)
        block_full(ws,r,1,r,16)
        put(ws,r,1,i+1,h='center')
        put(ws,r,2,line.hs_code_cn,size=9,h='center')
        elems = line.customs_elements
        if elems.startswith(line.customs_name_cn):
            nl = elems.find('\n'); elems = elems[nl+1:] if nl>=0 else elems
        dc = ws.cell(r,3, f"{line.customs_name_cn}\n{elems}")
        dc.font = F(False,8); dc.alignment = AL('left','top',True); dc.border = FULL()
        put(ws,r,5,f'{int(line.customs_suits)} 套',h='center')
        put(ws,r,7,ctnr_str,size=8,wrap=True)
        put(ws,r,8,f'{gw_val:,.0f}' if gw_val else '',h='center')
        put(ws,r,9,'泰国',h='center')
        put(ws,r,11,line.unit_price,h='center',num_fmt='#,##0.00')
        put(ws,r,12,line.total_amount,h='center',num_fmt='#,##0.00')
        put(ws,r,15,'USD',h='center'); put(ws,r,16,'照章',h='center')

    # 尾部：单位地址 / 申报单位签章 / 查验放行 / 邮编电话（录入员/录入单位已废止）
    tail = base_r + len(bundle.set1_lines)
    mc(ws,tail,1,tail,16)
    ac = ws.cell(tail,1, f'单位地址：{bundle.seller_address_en or ""}')
    ac.font = F(False,8); ac.alignment = AL('left','center',True); ac.border = FULL()
    ws.row_dimensions[tail].height = 18

    mc(ws,tail+1,1,tail+1,16)
    dc = ws.cell(tail+1,1,'兹申明对以上内容承担如实申报、依法纳税之法律责任')
    dc.font = F(False,8,GRAY,italic=True); dc.alignment = AL('left'); dc.border = FULL()
    ws.row_dimensions[tail+1].height = 14

    mc(ws,tail+2,1,tail+2,8); mc(ws,tail+2,9,tail+2,16)
    block_full(ws,tail+2,1,tail+2,16)
    put(ws,tail+2,1,'申报单位（签章）：',size=8)
    put(ws,tail+2,9,'海关批注及签章',size=8,h='center')
    ws.row_dimensions[tail+2].height = 28

    mc(ws,tail+3,1,tail+3,5); mc(ws,tail+3,6,tail+3,8)
    mc(ws,tail+3,9,tail+3,12); mc(ws,tail+3,13,tail+3,16)
    block_full(ws,tail+3,1,tail+3,16)
    put(ws,tail+3,1,'邮编/电话：',size=8)
    put(ws,tail+3,6,'填制日期：',size=8)
    put(ws,tail+3,9,bundle.invoice_date,size=8,h='center')
    put(ws,tail+3,13,'查验:        放行:',size=8,h='center')
    ws.row_dimensions[tail+3].height = 16

    setup_print(ws, orientation='landscape', header_row=HDR_R, fit_width_only=False)




def build_contract(ws, bundle):
    NC = 6
    for col,w in [('A',34),('B',10),('C',6),('D',16),('E',16),('F',20)]:
        ws.column_dimensions[col].width = w

    name_cn = bundle.seller_name_cn or bundle.seller_name_en

    # 标题
    mc_put(ws,1,1,1,NC,'SALES CONTRACT',bold=True,size=16,color=NAVY,h='center')
    ws.row_dimensions[1].height=26
    mc_put(ws,2,1,2,NC,'售货合同',bold=True,size=12,color=GRAY,h='center')
    ws.row_dimensions[2].height=18
    fill_border(ws,3,1,3,NC,t=False,b=True,l=False,r=False)
    ws.row_dimensions[3].height=4

    put(ws,4,1,f'合同编号 Contract No.: {bundle.invoice_no}',bold=True,size=10)
    put(ws,4,5,f'日期 Date: {bundle.contract_date}',bold=True,size=10,h='right')
    ws.row_dimensions[4].height=18; ws.row_dimensions[5].height=8

    # 卖方
    put(ws,6,1,'卖方 Seller:',bold=True,size=9,color=NAVY)
    ws.row_dimensions[6].height=14
    put(ws,7,1,f'{name_cn} / {bundle.seller_name_en}',bold=True,size=10)
    ws.row_dimensions[7].height=16
    put(ws,8,1,bundle.seller_address_en,size=9,color=GRAY)
    ws.row_dimensions[8].height=14
    ws.row_dimensions[9].height=8

    # 买方=货代
    put(ws,10,1,'买方 Buyer:',bold=True,size=9,color=NAVY)
    ws.row_dimensions[10].height=14
    put(ws,11,1,bundle.consignee_name,bold=True,size=10)
    ws.row_dimensions[11].height=16
    mc(ws,12,1,12,NC)
    put(ws,12,1,bundle.consignee_address,size=9,color=GRAY,wrap=True)
    ws.row_dimensions[12].height=20
    ws.row_dimensions[13].height=8

    mc(ws,14,1,14,NC)
    put(ws,14,1,'双方同意按下列条款由买方购进、卖方出售下列商品：',size=9)
    ws.row_dimensions[14].height=14
    mc(ws,15,1,15,NC)
    put(ws,15,1,'The Buyers agree to buy and the Sellers agree to sell the following goods on the terms stated below:',
        size=8,color=GRAY,italic=True)
    ws.row_dimensions[15].height=13
    ws.row_dimensions[16].height=6

    # 商品表头
    hdrs=[(1,'品名规格 Description'),(2,'数量 Qty'),(3,'单位'),
          (4,'单价 Unit Price\n(USD)'),(5,'总价 Amount\n(USD)'),(6,'')]
    mc(ws,17,5,17,6)
    for c,h in hdrs:
        if c==6: continue
        cell=ws.cell(17,c,h); cell.font=F(True,9,'FFFFFF'); cell.fill=FILL(NAVY)
        cell.border=FULL(); cell.alignment=AL('center')
    ws.row_dimensions[17].height=26

    r=18; total=0
    zebra_idx = 0
    for line in bundle.set1_lines:
        mc(ws,r,5,r,6)
        bg = ZEBRA if zebra_idx % 2 == 1 else None
        for c in range(1,7):
            ws.cell(r,c).border=FULL()
            if bg: ws.cell(r,c).fill=FILL(bg)
        put(ws,r,1,f"{line.customs_name_cn} / {line.name_en}",size=10,bg=bg)
        put(ws,r,2,int(line.customs_suits),h='right',num_fmt='#,##0',bg=bg)
        put(ws,r,3,'套',h='center',bg=bg)
        put(ws,r,4,line.unit_price,h='right',num_fmt='#,##0.00',bg=bg)
        put(ws,r,5,line.total_amount,h='right',num_fmt='#,##0.00',bg=bg)
        total += line.total_amount
        ws.row_dimensions[r].height=22; zebra_idx+=1; r+=1

    ws.row_dimensions[r].height=6; r+=1
    mc(ws,r,5,r,6)
    put(ws,r,1,'合计 TOTAL',bold=True,size=10,color=NAVY,h='right',border=FULL(),bg=LBLUE)
    for c in [2,3,4]: ws.cell(r,c).border=FULL(); ws.cell(r,c).fill=FILL(LBLUE)
    put(ws,r,5,total,bold=True,size=10,h='right',num_fmt='#,##0.00',border=FULL(),bg=LBLUE)
    ws.row_dimensions[r].height=22; r+=2

    terms = [
        f"1. 合同总值 Total Value: USD {total:,.2f}",
        "2. 包装 Packing: 纸箱 Carton",
        "3. 装运期限 Time of Shipment: （待安排）",
        f"4. 装运口岸 Port of Shipment: {bundle.seller_port_loading or 'NANSHA, China'}",
        f"5. 目的口岸 Port of Destination: {bundle.port_discharge or 'Bangkok, Thailand'}",
        f"6. 付款条件 Payment Terms: {bundle.buyer_payment_terms or 'T/T'}",
        f"7. 成交条件 Trade Terms: {bundle.buyer_incoterms or 'FOB'}",
        "8. 装运唛头 Shipping Marks: EF",
    ]
    for t in terms:
        mc(ws,r,1,r,NC); put(ws,r,1,t,size=9)
        ws.row_dimensions[r].height=16; r+=1
    ws.row_dimensions[r].height=10; r+=1

    fill_border(ws,r,1,r,NC,t=True,b=False,l=False,r=False)
    ws.row_dimensions[r].height=4; r+=1
    put(ws,r,1,'卖方 Seller:',bold=True,size=9,color=NAVY)
    put(ws,r,4,'买方 Buyer:',bold=True,size=9,color=NAVY)
    ws.row_dimensions[r].height=14; r+=1
    put(ws,r,1,name_cn,bold=True,size=10); ws.row_dimensions[r].height=16; r+=1
    put(ws,r,1,bundle.seller_name_en,size=9,color=GRAY)
    put(ws,r,4,bundle.consignee_name,bold=True,size=10)
    ws.row_dimensions[r].height=16


# ════════════════════════════════════════════════════════════
# SI (Shipping Instruction)
# ════════════════════════════════════════════════════════════
def build_si(ws, bundle):
    for col,w in [('A',20),('B',46),('C',14),('D',9),('E',9),
                   ('F',10),('G',10),('H',10),('I',10)]:
        ws.column_dimensions[col].width=w
    NC = 9

    so = '；'.join(dict.fromkeys(bundle.so_nos)) if bundle.so_nos else '（待录入）'
    mc_put(ws,1,1,1,NC, f'SO#: {so}', bold=True, size=11, color=NAVY)
    ws.row_dimensions[1].height=18
    ws.row_dimensions[2].height=6
    mc_put(ws,3,1,3,NC,'SHIPPING INSTRUCTION',bold=True,size=14,color=NAVY,h='center')
    ws.row_dimensions[3].height=22
    ws.row_dimensions[4].height=6

    r=5
    for lbl,content in [
        ('1. Shipper:', f'{bundle.seller_name_en}\n{bundle.seller_address_en}'),
        ('2. Consignee:', f'{bundle.consignee_name}\n{bundle.consignee_address}'),
        ('3. Notify Party:', f'{bundle.notify_name}\n{bundle.notify_address}'),
        ('4. Port of Loading:', bundle.seller_port_loading or 'NANSHA'),
        ('5. Port of Discharge:', bundle.port_discharge or 'Bangkok'),
    ]:
        put(ws,r,1,lbl,bold=True,size=10,color=NAVY)
        mc(ws,r,2,r,NC)
        put(ws,r,2,content,size=9,wrap=True)
        lines = content.count('\n')+1
        ws.row_dimensions[r].height = max(16, 16*lines)
        ws.row_dimensions[r+1].height = 5
        r += 2

    mc_put(ws,r,1,r,NC,'6. Description of Goods',bold=True,size=10,color=NAVY)
    ws.row_dimensions[r].height=16; r+=1
    for line in bundle.set1_lines:
        put(ws,r,1,f'• {line.name_en}',size=9)
        ws.row_dimensions[r].height=14; r+=1
    ws.row_dimensions[r].height=5; r+=1

    put(ws,r,1,'7. Shipping Marks: EF',bold=True,size=10,color=NAVY)
    ws.row_dimensions[r].height=16; r+=2

    # 货柜明细表
    hdrs = ['SO','Container No.','Seal No.','PKGS','Tare(kg)','G.W.(kg)','CBM','VGM(kg)','Type']
    for c,h in enumerate(hdrs,1):
        cell = ws.cell(r,c,h); cell.font=F(True,9,'FFFFFF'); cell.fill=FILL(NAVY)
        cell.border=FULL(); cell.alignment=AL('center')
    ws.row_dimensions[r].height=20; r+=1

    import math as _m
    tare_map = {'40HQ':3900,'45HQ':4200,'40GP':2300,'20GP':2200}
    tot_pkgs=0; tot_gw=0; tot_cbm=0
    zebra_idx = 0
    for ctnr in bundle.containers:
        rows = [l for l in bundle.set2_lines if l.container_seq==ctnr.seq]
        raw_pkgs = sum(int(getattr(l,'box_count',0)) for l in rows)
        if raw_pkgs == 0 and bundle.total_pkgs > 0:
            raw_pkgs = _m.ceil(bundle.total_pkgs / len(bundle.containers))
        gw = sum(l.total_gw for l in rows)
        cbm = sum(l.total_cbm for l in rows)
        sz = ctnr.container_size or '40HQ'
        tare = tare_map.get(sz, 3900)
        vgm = ctnr.vgm_kg if ctnr.vgm_kg else round(gw+tare,0)

        bg = ZEBRA if zebra_idx % 2 == 1 else None
        for c in range(1, NC+1):
            ws.cell(r,c).border = FULL()
            if bg: ws.cell(r,c).fill = FILL(bg)
        vals = [ctnr.so_no or '（待录入）', ctnr.container_no, ctnr.seal_no or '（待录入）',
                raw_pkgs, tare, round(gw,0) if gw else 0,
                round(cbm,1) if cbm else 0, int(vgm), sz]
        for c,v in enumerate(vals,1):
            num = c in (4,5,6,7,8)
            put(ws,r,c,v,size=9,h='right' if num else 'left',bg=bg)
        tot_pkgs+=raw_pkgs; tot_gw+=gw; tot_cbm+=cbm
        ws.row_dimensions[r].height=20; zebra_idx+=1; r+=1

    for c in range(1, NC+1): ws.cell(r,c).fill = FILL(LBLUE)
    block_full(ws, r, 1, r, NC)
    put(ws,r,1,'TOTAL',bold=True,bg=LBLUE)
    put(ws,r,4,tot_pkgs,bold=True,h='right',bg=LBLUE)
    put(ws,r,6,round(tot_gw,0) if tot_gw else 0,bold=True,h='right',bg=LBLUE)
    put(ws,r,7,round(tot_cbm,1) if tot_cbm else 0,bold=True,h='right',bg=LBLUE)
    ws.row_dimensions[r].height=20; r+=2

    mc(ws,r,1,r,NC)
    put(ws,r,1,'⚠ SO号、Seal No.、出口口岸等海关/货代专属信息需由单证员或货代补充录入',
        size=8, color=GRAY, italic=True)
    ws.row_dimensions[r].height=14


# ════════════════════════════════════════════════════════════
# 主入口
# ════════════════════════════════════════════════════════════
def generate_document_set(bundle, output_dir, master_path=None):
    os.makedirs(output_dir, exist_ok=True)
    date_str = bundle.invoice_date.replace('-','')
    base = f"{bundle.invoice_no}_{bundle.customer_code}_{date_str}"

    imgs = {}
    if master_path:
        try: imgs = load_component_images(master_path)
        except: pass

    # 套一
    wb1 = Workbook()
    ws = wb1.active; ws.title='Inv.'
    build_inv1(ws, bundle)
    build_pl1(wb1.create_sheet('PL'), bundle)
    build_customs(wb1.create_sheet('报关单'), bundle)
    build_contract(wb1.create_sheet('合同 '), bundle)
    build_si(wb1.create_sheet('SI'), bundle)
    path1 = os.path.join(output_dir, f'套一_{base}.xlsx')
    wb1.save(path1)

    # 套二
    wb2 = Workbook()
    ws2 = wb2.active; ws2.title='Inv.(Parts)'
    build_inv2(ws2, bundle, imgs)
    build_pl2(wb2.create_sheet('PL(Parts)'), bundle, imgs)
    path2 = os.path.join(output_dir, f'套二_{base}.xlsx')
    wb2.save(path2)

    return {'set1': path1, 'set2': path2}
